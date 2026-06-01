import base64
import csv
import difflib
import gzip
import hashlib
import json
import os
import re
import socket
import uuid
import urllib.request
from urllib.parse import quote
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from urllib.parse import urlencode

from django.conf import settings
from django.core.management import call_command
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from django.db import transaction
from django.db.models import Sum, F, Count, Q
from django.db.models.deletion import ProtectedError
from django.db.models.functions import TruncMonth
from django.core import serializers
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, HttpResponse, Http404, JsonResponse
from PIL import Image as PILImage
from django import forms
from django.contrib import messages
from django.utils import timezone
from django.urls import reverse

from .models import (
    Category, Product, Sale, SaleItem,
    Expense, Debt, Supplier, Promotion,
    Customer, PhoneCredit, PhoneCreditPurchase, StoreSettings,
    AppFeature, StockLoss, StockSupply, PurchasePriceHistory,
    PurchaseOrder, PurchaseOrderItem,
    ProductReturn, ProductReturnItem,
    CashSession, CashMovement,
    Recipe, RecipeIngredient,
    Employee, Attendance,
)

from .forms import (
    ProductForm, SaleForm, CartAddForm,
    ExpenseForm, DebtForm,
    PhoneCreditForm, PhoneCreditPurchaseForm,
    StoreSettingsForm, AppFeatureForm,
    AccountCreationForm, AIFeatureInstructionForm,
    DataImportForm, StockLossForm,
    StockSupplyForm,
    CashSessionOpenForm, CashSessionCloseForm, CashMovementForm,
    EmployeeCreateForm,
)


# =========================
# BARCODE GENERATION
# =========================
def _barcode_image(text, width=280, height=80):
    """Generate barcode PNG BytesIO from text using python-barcode."""
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
    buf = BytesIO()
    try:
        code = get_barcode_class('code128')
        writer = ImageWriter()
        writer.set_options({'module_width': 0.3, 'module_height': 10,
                            'font_size': 0, 'text_distance': 0,
                            'quiet_zone': 2})
        inst = code(text, writer=writer)
        inst.write(buf)
        buf.seek(0)
        img = PILImage.open(buf).convert('RGBA')
        img = img.resize((width, height), PILImage.LANCZOS)
        out = BytesIO()
        img.save(out, format='PNG')
        out.seek(0)
        return out
    except Exception:
        return None


def product_barcode_image(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if not product.barcode:
        raise Http404('Aucun code-barres')
    data = _barcode_image(product.barcode, 400, 120)
    if data is None:
        raise Http404('Erreur génération code-barres')
    return HttpResponse(data.getvalue(), content_type='image/png')


# =========================
# HOME
# =========================
def home(request):
    low_stock_count = Product.objects.filter(stock__lte=F('min_stock')).count()
    return render(request, 'store/home.html', {
        'low_stock_count': low_stock_count,
    })


# =========================
# DASHBOARD
# =========================
def dashboard_view(request):
    today = date.today()

    total_products = Product.objects.count()
    total_stock = Product.objects.aggregate(total=Sum('stock'))['total'] or 0
    total_sales = Sale.objects.count()
    total_revenue = Sale.objects.aggregate(total=Sum('total'))['total'] or 0
    daily_revenue = Sale.objects.filter(sale_date=today).aggregate(total=Sum('total'))['total'] or 0
    daily_count = Sale.objects.filter(sale_date=today).count()
    monthly_revenue = Sale.objects.filter(sale_date__year=today.year, sale_date__month=today.month).aggregate(total=Sum('total'))['total'] or 0
    prev_month = today.month - 1 or 12
    prev_year = today.year if today.month > 1 else today.year - 1
    prev_month_revenue = Sale.objects.filter(sale_date__year=prev_year, sale_date__month=prev_month).aggregate(total=Sum('total'))['total'] or 0
    month_change = ((monthly_revenue - prev_month_revenue) / prev_month_revenue * 100) if prev_month_revenue else 0
    total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    low_stock_count = Product.objects.filter(stock__lte=F('min_stock')).count()
    out_of_stock = Product.objects.filter(stock=0).count()
    total_cost = sum((item.product.cost_price * item.quantity)
                     for item in SaleItem.objects.select_related('product').all()) or 0
    total_profit = total_revenue - total_cost - total_expenses
    top_category = (SaleItem.objects.values('product__category__name')
                    .annotate(total=Sum('quantity'))
                    .order_by('-total').first())
    recent_sales = Sale.objects.order_by('-created_at')[:5]

    # AJAX refresh: return only the dynamic stats
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('refresh'):
        return JsonResponse({
            'monthly_revenue': float(monthly_revenue),
            'today_revenue': float(daily_revenue),
            'total_profit': float(total_profit),
            'total_expenses': float(total_expenses),
            'low_stock_count': low_stock_count,
            'out_of_stock': out_of_stock,
            'total_sales': total_sales,
            'daily_count': daily_count,
            'total_stock': total_stock,
        })

    daily_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        total = Sale.objects.filter(sale_date=day).aggregate(total=Sum('total'))['total'] or 0
        daily_data.append({'label': day.strftime('%d/%m'), 'value': float(total)})

    monthly_data = []
    for month in range(1, 13):
        total = Sale.objects.filter(sale_date__year=today.year, sale_date__month=month).aggregate(total=Sum('total'))['total'] or 0
        monthly_data.append({'label': date(1900, month, 1).strftime('%b'), 'value': float(total)})

    yearly_data = []
    for year in range(today.year - 2, today.year + 1):
        total = Sale.objects.filter(sale_date__year=year).aggregate(total=Sum('total'))['total'] or 0
        yearly_data.append({'label': str(year), 'value': float(total)})

    return render(request, 'store/dashboard.html', {
        'total_products': total_products,
        'total_stock': total_stock,
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'daily_revenue': daily_revenue,
        'daily_count': daily_count,
        'monthly_revenue': monthly_revenue,
        'prev_month_revenue': prev_month_revenue,
        'month_change': month_change,
        'total_expenses': total_expenses,
        'low_stock_count': low_stock_count,
        'out_of_stock': out_of_stock,
        'total_profit': total_profit,
        'top_category': top_category,
        'recent_sales': recent_sales,
        'monthly_expense_limit': float(StoreSettings.load().monthly_expense_limit or 0),
        'budget_pct': min(100, round((total_expenses / float(StoreSettings.load().monthly_expense_limit or 1)) * 100)) if StoreSettings.load().monthly_expense_limit else 0,
        'daily_labels': json.dumps([item['label'] for item in daily_data]),
        'daily_values': json.dumps([item['value'] for item in daily_data]),
        'month_labels': json.dumps([item['label'] for item in monthly_data]),
        'month_values': json.dumps([item['value'] for item in monthly_data]),
        'year_labels': json.dumps([item['label'] for item in yearly_data]),
        'year_values': json.dumps([item['value'] for item in yearly_data]),
    })


# =========================
# PRODUITS
# =========================
def product_list(request):
    form = CartAddForm()
    query = request.GET.get('q', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    sort = request.GET.get('sort', '')
    order = request.GET.get('order', 'asc')

    products = Product.objects.all()

    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(barcode__icontains=query) |
            Q(category__name__icontains=query)
        )
    try:
        if min_price:
            products = products.filter(price__gte=Decimal(min_price))
        if max_price:
            products = products.filter(price__lte=Decimal(max_price))
    except InvalidOperation:
        messages.warning(request, 'Filtre de prix invalide.')

    sort_fields = {
        'name': 'name',
        'price': 'price',
        'stock': 'stock',
        'category': 'category__name',
        'expiry': 'expiry_date',
    }
    if sort in sort_fields:
        order_prefix = '-' if order == 'desc' else ''
        products = products.order_by(f'{order_prefix}{sort_fields[sort]}')

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']

    return render(request, 'store/product_list.html', {
        'page_obj': page_obj,
        'form': form,
        'query': query,
        'min_price': min_price,
        'max_price': max_price,
        'current_sort': sort,
        'current_order': order,
        'query_params': qp.urlencode(),
    })


def price_labels(request):
    selected_category = request.GET.get('category', '')
    products = Product.objects.select_related('category').order_by('name')
    if selected_category:
        products = products.filter(category_id=selected_category)

    return render(request, 'store/price_labels.html', {
        'products': products,
        'categories': Category.objects.order_by('name'),
        'selected_category': selected_category,
    })


def price_labels_pdf(request):
    selected_category = request.GET.get('category', '')
    products = Product.objects.select_related('category').order_by('name')
    if selected_category:
        products = products.filter(category_id=selected_category)

    settings_obj = StoreSettings.load()
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    margin = 18
    columns = max(2, min(settings_obj.label_columns_print or 4, 8))
    rows = 5
    gap = 6
    card_w = (width - (margin * 2) - (gap * (columns - 1))) / columns
    card_h = (height - (margin * 2) - (gap * (rows - 1))) / rows

    for index, product in enumerate(products):
        slot = index % (columns * rows)
        if index and slot == 0:
            pdf.showPage()
        row = slot // columns
        col = slot % columns
        x = margin + col * (card_w + gap)
        y = height - margin - card_h - row * (card_h + gap)

        if settings_obj.label_border_style != 'none':
            if settings_obj.label_border_style == 'dashed':
                pdf.setDash(3, 2)
            elif settings_obj.label_border_style == 'dotted':
                pdf.setDash(1, 2)
            else:
                pdf.setDash()
            pdf.setStrokeColorRGB(.65, .65, .65)
            pdf.rect(x, y, card_w, card_h, stroke=1, fill=0)
            pdf.setDash()

        text_y = y + card_h - 14
        pdf.setFillColorRGB(.08, .1, .13)
        if settings_obj.label_show_store_name:
            pdf.setFont('Helvetica-Bold', 6)
            pdf.drawCentredString(x + card_w / 2, text_y, settings_obj.store_name[:38])
            text_y -= 11

        pdf.setFont('Helvetica-Bold', 8 if settings_obj.label_font_size != 'large' else 10)
        pdf.drawCentredString(x + card_w / 2, text_y, product.name[:34])
        text_y -= 16

        pdf.setFont('Helvetica-Bold', 15 if settings_obj.label_font_size != 'small' else 12)
        pdf.drawCentredString(x + card_w / 2, text_y, f'{product.price:.0f} FCFA')
        text_y -= 12

        pdf.setFont('Helvetica', 6)
        pdf.drawCentredString(x + card_w / 2, text_y, product.get_unit_display())
        text_y -= 9
        if product.barcode and settings_obj.label_show_barcode:
            barcode_data = _barcode_image(str(product.barcode)[:30], int(card_w - 10), 28)
            if barcode_data:
                try:
                    img = PILImage.open(barcode_data)
                    tmp = BytesIO()
                    img.save(tmp, format='PNG')
                    tmp.seek(0)
                    barcode_x = x + (card_w - (card_w - 10)) / 2
                    barcode_y = text_y - 28
                    pdf.drawInlineImage(tmp, barcode_x, barcode_y, width=card_w - 10, height=28)
                    text_y = barcode_y - 2
                except Exception:
                    pdf.drawCentredString(x + card_w / 2, text_y, str(product.barcode)[:42])
                    text_y -= 8
            else:
                pdf.drawCentredString(x + card_w / 2, text_y, str(product.barcode)[:42])
                text_y -= 8
        if product.code and settings_obj.label_show_code:
            pdf.drawCentredString(x + card_w / 2, text_y, str(product.code)[:34])

    pdf.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=False, filename='etiquettes-prix.pdf')


def store_settings_view(request):
    settings = StoreSettings.load()
    form = StoreSettingsForm(request.POST or None, request.FILES or None, instance=settings)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Paramètres enregistrés avec succès.')
        if request.GET.get('embed') == '1':
            return redirect(f"{reverse('store:store_settings')}?embed=1&saved=1")
        return redirect('store:store_settings')

    return render(request, 'store/settings_form.html', {'form': form})


def connected_devices(request):
    from django.contrib.sessions.models import Session
    from django.utils import timezone
    now = timezone.now()
    sessions = Session.objects.filter(expire_date__gte=now)
    devices = []
    for s in sessions:
        try:
            data = s.get_decoded()
            devices.append({
                'session_key': s.session_key,
                'ip': data.get('ip_address', 'Inconnu'),
                'user_agent': data.get('user_agent', 'Inconnu'),
                'last_activity': data.get('last_activity', s.expire_date),
                'expire_date': s.expire_date,
            })
        except Exception:
            pass
    devices.sort(key=lambda d: d['last_activity'], reverse=True)
    return render(request, 'store/connected_devices.html', {'devices': devices})


def disconnect_device(request, session_key):
    from django.contrib.sessions.models import Session
    try:
        Session.objects.filter(session_key=session_key).delete()
        messages.success(request, 'Appareil déconnecté.')
    except Exception:
        messages.error(request, 'Erreur lors de la déconnexion.')
    return redirect('store:connected_devices')


def mobile_access(request):
    hostnames = []
    try:
        hostname = socket.gethostname()
        for item in socket.getaddrinfo(hostname, None):
            ip = item[4][0]
            if '.' in ip and not ip.startswith('127.') and ip not in hostnames:
                hostnames.append(ip)
    except OSError:
        pass

    return render(request, 'store/mobile_access.html', {'local_ips': hostnames})


# =========================
# CAISSE JOURNALIÈRE
# =========================
def cash_session_list(request):
    sessions = CashSession.objects.all()
    active_session = sessions.filter(closed_at__isnull=True).first()
    open_form = CashSessionOpenForm()
    close_form = CashSessionCloseForm()
    movement_form = CashMovementForm()
    if active_session:
        close_form = CashSessionCloseForm(instance=active_session)
        movement_form = CashMovementForm(initial={'session': active_session.pk})
    return render(request, 'store/cash_session_list.html', {
        'sessions': sessions,
        'active_session': active_session,
        'open_form': open_form,
        'close_form': close_form,
        'movement_form': movement_form,
    })


def cash_session_open(request):
    if request.method == 'POST':
        form = CashSessionOpenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Caisse ouverte.')
        else:
            messages.error(request, 'Erreur à l\'ouverture.')
    return redirect('store:cash_session_list')


def cash_session_close(request, pk):
    session = get_object_or_404(CashSession, pk=pk)
    if request.method == 'POST':
        form = CashSessionCloseForm(request.POST, instance=session)
        if form.is_valid():
            session = form.save(commit=False)
            session.closed_at = timezone.now()
            session.save()
            messages.success(request, 'Caisse clôturée.')
        else:
            messages.error(request, 'Erreur à la clôture.')
    return redirect('store:cash_session_list')


def cash_session_restore(request, pk):
    session = get_object_or_404(CashSession, pk=pk)
    if session.closed_at:
        session.closed_at = None
        session.closing_balance = None
        session.save()
        messages.success(request, 'Session caisse restaurée.')
    return redirect('store:cash_session_list')


def cash_movement_create(request):
    if request.method == 'POST':
        session_pk = request.POST.get('session')
        session = get_object_or_404(CashSession, pk=session_pk) if session_pk else None
        if session:
            form = CashMovementForm(request.POST)
            if form.is_valid():
                movement = form.save(commit=False)
                movement.session = session
                movement.save()
                messages.success(request, 'Mouvement ajouté.')
            else:
                messages.error(request, 'Erreur mouvement.')
    return redirect('store:cash_session_list')


def database_tools(request):
    db_config = settings.DATABASES['default']
    db_engine = 'postgresql' if 'postgresql' in db_config.get('ENGINE', '') else 'sqlite'
    db_name = db_config.get('NAME') or db_config.get('HOST', '')
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    backups = sorted(
        [f for f in os.listdir(backup_dir) if f.endswith('.json.gz') or f.endswith('.bak')],
        reverse=True
    )[:10]
    return render(request, 'store/database_tools.html', {
        'db_engine': db_engine,
        'db_name': db_name,
        'backups': backups,
        'backup_dir': backup_dir,
    })


def _encrypt(data, password):
    if not password:
        return data
    key = hashlib.sha256(password.encode()).digest()
    encrypted = bytearray()
    for i, b in enumerate(data):
        encrypted.append(b ^ key[i % len(key)])
    return base64.b64encode(bytes(encrypted))


def _decrypt(data, password):
    if not password:
        return data
    key = hashlib.sha256(password.encode()).digest()
    raw = base64.b64decode(data)
    decrypted = bytearray()
    for i, b in enumerate(raw):
        decrypted.append(b ^ key[i % len(key)])
    return bytes(decrypted)


def database_backup(request):
    password = request.POST.get('backup_password', '')
    buf = StringIO()
    call_command('dumpdata', 'store', indent=2, stdout=buf, exclude=['contenttypes', 'auth.permission'])
    raw = buf.getvalue().encode('utf-8')
    compressed = gzip.compress(raw)
    if password:
        compressed = _encrypt(compressed, password)
        filename = 'sauvegarde_supermarche.bak'
        content_type = 'application/octet-stream'
    else:
        filename = 'sauvegarde_supermarche.json.gz'
        content_type = 'application/gzip'
    response = HttpResponse(compressed, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def database_export_json(request):
    response = HttpResponse(content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="donnees_supermarche.json"'
    call_command(
        'dumpdata',
        'store',
        indent=2,
        stdout=response,
        exclude=['contenttypes', 'auth.permission'],
    )
    return response


def database_import_json(request):
    form = DataImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        uploaded_file = request.FILES['file']
        if not uploaded_file.name.endswith('.json'):
            messages.error(request, 'Le fichier doit etre au format JSON.')
    return redirect('store:database_tools')


# ===== RAPPORT JOURNALIER =====
def daily_report_view(request):
    today = date.today()
    sales_today = Sale.objects.filter(sale_date=today)
    daily_total = sales_today.aggregate(total=Sum('total'))['total'] or 0
    daily_count = sales_today.count()
    daily_cost = sum(
        (item.product.cost_price or 0) * item.quantity
        for sale in sales_today.prefetch_related('items__product')
        for item in sale.items.all()
    )
    daily_profit = daily_total - daily_cost
    expenses_today = Expense.objects.filter(date=today).aggregate(total=Sum('amount'))['total'] or 0
    top_products = (
        SaleItem.objects.filter(sale__sale_date=today)
        .values('product__name')
        .annotate(total_qty=Sum('quantity'), total_rev=Sum('price'))
        .order_by('-total_qty')[:5]
    )
    payment_breakdown = (
        sales_today.values('payment_method')
        .annotate(total=Sum('total'), count=Count('id'))
    )
    cash_sessions = CashSession.objects.filter(opened_at__date=today)
    return render(request, 'store/daily_report.html', {
        'daily_total': daily_total,
        'daily_count': daily_count,
        'daily_profit': daily_profit,
        'daily_cost': daily_cost,
        'expenses_today': expenses_today,
        'top_products': top_products,
        'payment_breakdown': payment_breakdown,
        'cash_sessions': cash_sessions,
    })


# ===== INVENTAIRE RAPIDE =====
def quick_inventory(request):
    products = Product.objects.select_related('category').order_by('name')
    if request.method == 'POST':
        counts = {}
        for key, val in request.POST.items():
            if key.startswith('qty_'):
                try:
                    pk = int(key.replace('qty_', ''))
                    qty = Decimal(val)
                    product = Product.objects.get(pk=pk)
                    diff = qty - product.stock
                    InventoryAdjustment.objects.create(
                        product=product, system_stock=product.stock,
                        counted_stock=qty, difference=diff,
                        notes='Inventaire rapide',
                    )
                    product.stock = qty
                    product.save(update_fields=['stock'])
                except Exception:
                    pass
        messages.success(request, 'Inventaire terminé.')
        return redirect('store:inventory_list')
    return render(request, 'store/quick_inventory.html', {'products': products})


# ===== HISTORIQUE MOUVEMENTS STOCK =====
def stock_movement_history(request):
    from .models import ActivityLog
    movements = ActivityLog.objects.filter(
        model_name__in=['Product', 'StockSupply', 'InventoryAdjustment', 'StockLoss']
    ).order_by('-created_at')[:100]
    return render(request, 'store/stock_movement_history.html', {'movements': movements})


# ===== MARGE DU JOUR =====
def daily_margin_data(request):
    today = date.today()
    sales_today = Sale.objects.filter(sale_date=today)
    total = float(sales_today.aggregate(total=Sum('total'))['total'] or 0)
    cost = sum(
        float((item.product.cost_price or 0) * item.quantity)
        for sale in sales_today.prefetch_related('items__product')
        for item in sale.items.all()
    )
    expenses = float(Expense.objects.filter(date=today).aggregate(total=Sum('amount'))['total'] or 0)
    profit = total - cost - expenses
    margin_pct = (profit / total * 100) if total else 0
    product_margins = []
    for item in SaleItem.objects.filter(sale__sale_date=today).select_related('product').values('product__name').annotate(
        total_rev=Sum('price'), total_cost=Sum(F('product__cost_price') * F('quantity'))
    ):
        product_margins.append(item)
    return JsonResponse({
        'total': total, 'cost': cost, 'expenses': expenses,
        'profit': profit, 'margin_pct': round(margin_pct, 1),
        'product_margins': list(product_margins),
    })


# ===== SAUVEGARDE PROGRAMMÉE =====
def scheduled_backup(request):
    import shutil, datetime, glob
    db_path = None
    for alias in settings.DATABASES:
        opt = settings.DATABASES[alias]
        if 'NAME' in opt and str(opt['NAME']).endswith('.sqlite3'):
            db_path = opt['NAME']
            break
    if not db_path:
        return JsonResponse({'error': 'DB not found'}, status=400)
    backup_dir = os.path.join(os.path.dirname(db_path), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_{ts}.sqlite3'
    shutil.copy2(db_path, os.path.join(backup_dir, filename))
    # Rotation: keep last 20
    backups = sorted(glob.glob(os.path.join(backup_dir, 'backup_*.sqlite3')))
    for old in backups[:-20]:
        os.remove(old)
    return JsonResponse({'backup': filename, 'kept': len(backups[-20:])})


# ===== NETTOYAGE RAPIDE =====
def cleanup_old_data(request):
    from datetime import timedelta
    cutoff = date.today() - timedelta(days=365)
    deleted_sales = Sale.objects.filter(sale_date__lt=cutoff).count()
    Sale.objects.filter(sale_date__lt=cutoff).delete()
    deleted_returns = ProductReturn.objects.filter(date__lt=cutoff).count()
    ProductReturn.objects.filter(date__lt=cutoff).delete()
    messages.success(request, f'Nettoyage : {deleted_sales} vente(s) et {deleted_returns} retour(s) supprimés.')
    return redirect('store:database_tools')


# ===== ALERTE RUPTURE SMS FOURNISSEUR =====
def alert_supplier_rupture(request, pk):
    product = get_object_or_404(Product, pk=pk)
    supplies = StockSupply.objects.filter(product=product).select_related('supplier').order_by('-date')
    supplier = supplies.first().supplier if supplies.exists() else None
    if not supplier or not supplier.phone:
        messages.error(request, 'Aucun fournisseur avec téléphone pour ce produit.')
        return redirect('store:product_manage')
    msg = f'URGENT: {product.name} en rupture. Besoin de réapprovisionnement urgent.'
    try:
        from .notifications import send_sms
        send_sms(supplier.phone, msg)
        messages.success(request, f'SMS envoyé à {supplier.name} ({supplier.phone}).')
    except Exception as e:
        messages.error(request, f'Erreur envoi SMS: {e}')
    return redirect('store:out_of_stock')


# ===== RÉSUMÉ RACCOURCIS CLAVIER =====
def keyboard_shortcuts_help(request):
    shortcuts = [
        ('Alt+1', 'Accueil'), ('Alt+2', 'Caisse tactile'), ('Alt+3', 'Produits'),
        ('Ctrl+K', 'Recherche rapide'), ('Ctrl+F', 'Recherche globale'),
        ('N', 'Nouvelle vente'), ('F2', 'Rechercher (caisse)'),
        ('?', 'Aide raccourcis'), ('Escape', 'Fermer modale'),
    ]
    return render(request, 'store/keyboard_shortcuts.html', {'shortcuts': shortcuts})


# ===== RETOURS / ÉCHANGES =====
def return_list(request):
    returns = ProductReturn.objects.select_related('sale').all()
    paginator = Paginator(returns, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'store/return_list.html', {'page_obj': page_obj})


def return_create(request):
    from django.forms import inlineformset_factory
    ReturnItemFormSet = inlineformset_factory(
        ProductReturn, ProductReturnItem,
        fields=['product', 'quantity', 'refund_amount'],
        extra=3, can_delete=True,
        widgets={
            'product': forms.Select(attrs={'class': 'form-select form-select-sm'}),
            'quantity': forms.NumberInput(attrs={'class': 'form-control form-control-sm', 'step': '0.01', 'min': '0'}),
            'refund_amount': forms.NumberInput(attrs={'class': 'form-control form-control-sm', 'step': '0.01', 'min': '0'}),
        }
    )
    if request.method == 'POST':
        form = ProductReturnForm(request.POST)
        if form.is_valid():
            ret = form.save()
            formset = ReturnItemFormSet(request.POST, instance=ret)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Retour enregistré.')
                return redirect('store:return_list')
            else:
                ret.delete()
    else:
        form = ProductReturnForm()
        formset = ReturnItemFormSet()
    return render(request, 'store/return_form.html', {'form': form, 'formset': formset})


@transaction.atomic
def return_approve(request, pk):
    ret = get_object_or_404(ProductReturn, pk=pk)
    if ret.status != 'pending':
        messages.warning(request, f'Ce retour est déjà {ret.get_status_display().lower()}.')
        return redirect('store:return_list')
    ret.status = 'approved'
    ret.save()
    restocked = 0
    for item in ret.items.select_related('product'):
        product = item.product
        qty = float(item.quantity)
        if qty > 0:
            product.stock += qty
            product.save(update_fields=['stock'])
            restocked += 1
    messages.success(request, f'Retour #{ret.id} approuvé et {restocked} produit(s) remis en stock.')
    return redirect('store:return_list')


def return_reject(request, pk):
    ret = get_object_or_404(ProductReturn, pk=pk)
    if ret.status != 'pending':
        messages.warning(request, f'Ce retour est déjà {ret.get_status_display().lower()}.')
        return redirect('store:return_list')
    ret.status = 'rejected'
    ret.save()
    messages.info(request, f'Retour #{ret.id} rejeté.')
    return redirect('store:return_list')


# ===== RÉAPPROVISIONNEMENT AUTO =====
def auto_reorder(request):
    from django.db.models import Sum, F
    from datetime import timedelta
    today = date.today()
    products_to_order = []
    products = Product.objects.filter(stock__lte=F('min_stock')).order_by('stock')
    for p in products:
        last_30_days = today - timedelta(days=30)
        sales_qty = SaleItem.objects.filter(
            product=p, sale__sale_date__gte=last_30_days
        ).aggregate(total=Sum('quantity'))['total'] or 0
        sale_velocity = round(sales_qty / 30, 1)
        days_remaining = int(p.stock / sale_velocity) if sale_velocity > 0 else 999
        suggested = max(p.min_stock * 2 - p.stock, 0)
        if suggested > 0:
            products_to_order.append({
                'product': p,
                'stock': p.stock,
                'min_stock': p.min_stock,
                'sale_velocity': sale_velocity,
                'days_remaining': days_remaining,
                'suggested_qty': suggested,
            })
    return render(request, 'store/auto_reorder.html', {'products': products_to_order})


def feature_list(request):
    features = AppFeature.objects.all()
    counts = {
        'idea': features.filter(status='idea').count(),
        'planned': features.filter(status='planned').count(),
        'coding': features.filter(status='coding').count(),
        'done': features.filter(status='done').count(),
    }
    paginator = Paginator(features, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']
    return render(request, 'store/feature_list.html', {
        'page_obj': page_obj,
        'counts': counts,
        'query_params': qp.urlencode(),
    })


def feature_create(request):
    form = AppFeatureForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Fonctionnalité ajoutée avec succès.')
        return redirect('store:feature_list')
    return render(request, 'store/feature_form.html', {'form': form, 'title': 'Nouvelle fonctionnalité'})


def ai_feature_assistant(request):
    form = AIFeatureInstructionForm(request.POST or None)
    suggestion = ''

    if request.method == 'POST' and form.is_valid():
        instruction = form.cleaned_data['instruction'].strip()
        try:
            from openai import OpenAI
            client = OpenAI()
            response = client.responses.create(
                model=settings.OPENAI_MODEL,
                input=(
                    "Tu aides a modifier une application Django de gestion de magasin. "
                    "Reponds en francais avec un titre court, les fichiers a modifier, "
                    "les etapes de code, les validations a faire, et les risques. "
                    "Ne dis jamais que la modification est deja faite. Instruction: "
                    f"{instruction}"
                ),
            )
            suggestion = response.output_text
        except Exception as exc:
            suggestion = (
                "IA indisponible pour le moment. Verifie que la dependance openai est installee "
                "et que la variable OPENAI_API_KEY est configuree.\n\n"
                f"Erreur technique: {exc}"
            )

    return render(request, 'store/ai_feature_assistant.html', {'form': form, 'suggestion': suggestion})


def feature_update(request, pk):
    feature = get_object_or_404(AppFeature, pk=pk)
    form = AppFeatureForm(request.POST or None, instance=feature)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Fonctionnalité mise à jour.')
        return redirect('store:feature_list')
    return render(request, 'store/feature_form.html', {'form': form, 'title': 'Modifier la fonctionnalité'})


def feature_delete(request, pk):
    feature = get_object_or_404(AppFeature, pk=pk)
    if request.method == 'POST':
        feature.delete()
        messages.success(request, 'Fonctionnalité supprimée.')
        return redirect('store:feature_list')
    return render(request, 'store/feature_confirm_delete.html', {'feature': feature})


def modification_guide(request):
    return render(request, 'store/modification_guide.html')


def account_create(request):
    form = AccountCreationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Compte utilisateur créé avec succès.')
        return redirect('store:store_settings')
    return render(request, 'store/account_form.html', {'form': form})


def product_manage(request):
    sort = request.GET.get('sort', '')
    order = request.GET.get('order', 'asc')
    query = request.GET.get('q', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')

    products = Product.objects.select_related('category')

    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(barcode__icontains=query) |
            Q(category__name__icontains=query)
        )
    try:
        if min_price:
            products = products.filter(price__gte=Decimal(min_price))
        if max_price:
            products = products.filter(price__lte=Decimal(max_price))
    except InvalidOperation:
        messages.warning(request, 'Filtre de prix invalide.')

    sort_fields = {
        'name': 'name',
        'category': 'category__name',
        'price': 'price',
        'cost_price': 'cost_price',
        'stock': 'stock',
        'expiry': 'expiry_date',
    }
    if sort in sort_fields:
        order_prefix = '-' if order == 'desc' else ''
        products = products.order_by(f'{order_prefix}{sort_fields[sort]}')
    else:
        products = products.order_by('name')
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']
    search_params = qp.copy()
    for key in ('sort', 'order'):
        search_params.pop(key, None)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    template_name = 'store/_product_manage_ajax.html' if is_ajax else 'store/product_manage.html'
    return render(request, template_name, {
        'page_obj': page_obj,
        'current_sort': sort,
        'current_order': order,
        'query_params': qp.urlencode(),
        'search_params': search_params.urlencode(),
        'query': query,
        'min_price': min_price,
        'max_price': max_price,
        'categories': Category.objects.all(),
    })


def download_image_from_url(url):
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            content = response.read()
        ext = os.path.splitext(urllib.parse.urlparse(url).path)[1] or '.jpg'
        filename = f'{uuid.uuid4().hex}{ext}'
        from django.core.files.base import ContentFile
        return ContentFile(content, name=filename)
    except Exception:
        return None


def product_create(request):
    form = ProductForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        product = form.save(commit=False)
        if not request.FILES.get('image') and form.cleaned_data.get('image_url'):
            img_file = download_image_from_url(form.cleaned_data['image_url'])
            if img_file:
                product.image.save(*img_file)
        product.save()
        messages.success(request, 'Produit ajouté avec succès.')
        return redirect('store:product_manage')
    return render(request, 'store/product_form.html', {'form': form, 'title': 'Nouveau produit'})


def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    old_values = {
        'name': product.name,
        'price': str(product.price),
        'cost_price': str(product.cost_price),
        'stock': str(product.stock),
        'min_stock': str(product.min_stock),
        'category': str(product.category) if product.category else '',
        'unit': product.unit,
    }
    form = ProductForm(request.POST or None, request.FILES or None, instance=product)
    if request.method == 'POST' and form.is_valid():
        product = form.save(commit=False)
        if not request.FILES.get('image') and form.cleaned_data.get('image_url'):
            img_file = download_image_from_url(form.cleaned_data['image_url'])
            if img_file:
                product.image.save(*img_file)
        product.save()
        if str(product.cost_price) != old_values['cost_price']:
            PurchasePriceHistory.objects.create(
                product=product,
                old_price=old_values['cost_price'],
                new_price=product.cost_price,
                source='manual',
            )
        changes = []
        if product.name != old_values['name']:
            changes.append(f"nom: {old_values['name']} → {product.name}")
        if str(product.price) != old_values['price']:
            changes.append(f"prix: {old_values['price']} → {product.price}")
        if str(product.cost_price) != old_values['cost_price']:
            changes.append(f"prix achat: {old_values['cost_price']} → {product.cost_price}")
        if str(product.stock) != old_values['stock']:
            changes.append(f"stock: {old_values['stock']} → {product.stock}")
        if str(product.min_stock) != old_values['min_stock']:
            changes.append(f"stock min: {old_values['min_stock']} → {product.min_stock}")
        new_category = str(product.category) if product.category else ''
        if new_category != old_values['category']:
            changes.append(f"catégorie: {old_values['category'] or '—'} → {new_category or '—'}")
        if product.unit != old_values['unit']:
            changes.append(f"unité: {old_values['unit']} → {product.unit}")
        msg = '✅ Produit modifié avec succès.'
        if changes:
            msg += '<br><small>' + '<br>'.join(changes[:5]) + '</small>'
            if len(changes) > 5:
                msg += f'<br><small>…et {len(changes) - 5} autre(s) modification(s)</small>'
        messages.success(request, msg)
        return redirect('store:product_manage')
    return render(request, 'store/product_form.html', {'form': form, 'title': 'Modifier le produit'})


def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        try:
            product.delete()
            messages.success(request, 'Produit supprimé avec succès.')
        except ProtectedError:
            messages.error(
                request,
                "Ce produit a déjà été vendu. Modifie ou supprime d'abord la vente concernée."
            )
        return redirect('store:product_manage')
    return render(request, 'store/product_confirm_delete.html', {'product': product})


def product_supply(request, pk):
    product = get_object_or_404(Product, pk=pk)
    old_cost_price = product.cost_price
    form = StockSupplyForm(request.POST or None, product=product)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            total_units = form.calculated_total_units()
            supply = form.save(commit=False)
            supply.product = product
            supply.total_units = total_units
            supply.total_cost = total_units * supply.unit_cost_price
            supply.save()
            product.pack_size = supply.units_per_package
            product.carton_size = supply.packages_per_carton
            product.cost_price = supply.unit_cost_price
            product.price = supply.unit_sale_price
            product.stock += total_units
            if supply.unit_cost_price:
                product.target_margin_percent = ((supply.unit_sale_price - supply.unit_cost_price) / supply.unit_cost_price) * 100
            product.save(update_fields=['pack_size', 'carton_size', 'cost_price', 'price', 'stock', 'target_margin_percent'])
        if supply.unit_cost_price != old_cost_price:
            PurchasePriceHistory.objects.create(
                product=product,
                old_price=old_cost_price or 0,
                new_price=supply.unit_cost_price,
                source='supply',
            )
        messages.success(request, f'Approvisionnement validé : {total_units} unité(s) ajoutée(s) au stock.')
        return redirect('store:product_manage')
    return render(request, 'store/product_supply_form.html', {
        'form': form,
        'product': product,
    })


def price_history(request, pk):
    product = get_object_or_404(Product, pk=pk)
    history = PurchasePriceHistory.objects.filter(product=product).select_related('product')
    paginator = Paginator(history, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'store/price_history.html', {
        'product': product,
        'page_obj': page_obj,
    })


def product_quick_stock(request, pk):
    if request.method != 'POST':
        return redirect('store:product_manage')
    product = get_object_or_404(Product, pk=pk)
    try:
        qty = Decimal(str(request.POST.get('qty', '0')))
    except Exception:
        qty = Decimal('0')
    if qty <= 0:
        messages.warning(request, f'Quantité invalide pour {product.name}')
        return redirect('store:product_manage')
    product.stock += qty
    cost_price = request.POST.get('cost_price', '').strip()
    if cost_price:
        try:
            product.cost_price = Decimal(cost_price)
        except Exception:
            pass
    product.save(update_fields=['stock', 'cost_price'])
    messages.success(request, f'+{qty} {product.get_unit_display()} ajouté(s) au stock de {product.name}')
    return redirect('store:product_manage')


def supply_history(request):
    supplies = StockSupply.objects.select_related('product', 'supplier').all()
    paginator = Paginator(supplies, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'store/supply_history.html', {'page_obj': page_obj})


def reception_list(request):
    pending_orders = PurchaseOrder.objects.filter(status__in=['sent', 'partial']).prefetch_related('items__product', 'supplier')
    recent_supplies = StockSupply.objects.select_related('product', 'supplier').order_by('-date', '-created_at')[:20]
    products = Product.objects.filter(active=True).order_by('name')
    form = StockSupplyForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            product = get_object_or_404(Product, pk=request.POST.get('product'))
            total_units = form.calculated_total_units()
            supply = form.save(commit=False)
            supply.product = product
            supply.total_units = total_units
            supply.total_cost = total_units * supply.unit_cost_price
            supply.save()
            product.pack_size = supply.units_per_package
            product.carton_size = supply.packages_per_carton
            product.cost_price = supply.unit_cost_price
            product.price = supply.unit_sale_price
            product.stock += total_units
            if supply.unit_cost_price:
                product.target_margin_percent = ((supply.unit_sale_price - supply.unit_cost_price) / supply.unit_cost_price) * 100
            product.save(update_fields=['pack_size', 'carton_size', 'cost_price', 'price', 'stock', 'target_margin_percent'])
            if supply.unit_cost_price:
                PurchasePriceHistory.objects.create(
                    product=product, old_price=product.cost_price or 0,
                    new_price=supply.unit_cost_price, source='supply',
                )
            messages.success(request, f'Réception validée : {total_units} unité(s) de {product.name} ajoutée(s) au stock.')
        return redirect('store:reception_list')
    return render(request, 'store/reception_list.html', {
        'pending_orders': pending_orders,
        'form': form,
        'products': products,
        'recent_supplies': recent_supplies,
    })


def reception_receive_order(request, pk):
    order = get_object_or_404(PurchaseOrder.objects.prefetch_related('items__product', 'supplier'), pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            for item in order.items.all():
                qty = request.POST.get(f'received_{item.pk}', item.remaining)
                try:
                    qty = Decimal(str(qty))
                    qty = max(0, min(qty, item.remaining))
                except (InvalidOperation, TypeError):
                    qty = 0
                if qty > 0:
                    item.quantity_received += qty
                    item.save(update_fields=['quantity_received'])
                    supply = StockSupply.objects.create(
                        product=item.product,
                        supplier=order.supplier,
                        quantity=qty,
                        supply_mode='piece',
                        total_units=qty,
                        unit_cost_price=item.unit_price,
                        unit_sale_price=item.product.price,
                        total_cost=qty * item.unit_price,
                        notes=f'Réception commande #{order.id}',
                    )
                    item.product.stock += qty
                    item.product.save(update_fields=['stock'])
                    old_cost = item.product.cost_price or 0
                    if item.unit_price != old_cost:
                        PurchasePriceHistory.objects.create(
                            product=item.product, old_price=old_cost,
                            new_price=item.unit_price, source='supply',
                        )
            all_received = all(item.remaining == 0 for item in order.items.all())
            order.status = 'received' if all_received else 'partial'
            order.save(update_fields=['status'])
        messages.success(request, f'Commande #{order.id} réceptionnée avec succès.')
        return redirect('store:reception_list')
    return render(request, 'store/purchase_order_receive.html', {'order': order})


def create_sale(request):
    products = Product.objects.select_related('category').all()
    quick_products = (
        Product.objects.filter(name__icontains='pain')
        | Product.objects.filter(name__icontains='glace')
        | Product.objects.filter(category__name__icontains='pain')
        | Product.objects.filter(category__name__icontains='glace')
    ).distinct()[:8]

    if request.method == 'POST':
        product_ids = request.POST.getlist('product')
        quantities = request.POST.getlist('quantity')
        sale_modes = request.POST.getlist('sale_mode')
        sale_date_str = request.POST.get('sale_date', '')
        sale_notes = request.POST.get('notes', '')
        customer_name = request.POST.get('customer_name', '').strip()
        customer_phone = request.POST.get('customer_phone', '').strip()
        rows = []
        errors = []

        try:
            sale_date = date.fromisoformat(sale_date_str) if sale_date_str else date.today()
        except (ValueError, TypeError):
            sale_date = date.today()

        for index, product_id in enumerate(product_ids):
            quantity_value = quantities[index] if index < len(quantities) else ''
            mode = sale_modes[index] if index < len(sale_modes) else 'piece'
            if not product_id and not quantity_value:
                continue
            try:
                product = Product.objects.get(pk=product_id)
                quantity = Decimal(quantity_value)
                if quantity <= 0:
                    raise InvalidOperation
            except (Product.DoesNotExist, InvalidOperation, ValueError):
                errors.append('Une ligne de vente est invalide.')
                continue

            PACK_MODES = {'paquet': 1, 'carton': 12, 'cartouche': 1}
            if mode in PACK_MODES and product.pack_size > 1:
                multiplier = PACK_MODES[mode]
                effective_qty = quantity * product.pack_size * multiplier
            else:
                effective_qty = quantity
                if mode not in ('kg', 'l'):
                    mode = 'piece'

            rows.append((product, quantity, mode, effective_qty))

        if not rows:
            errors.append('Ajoute au moins un produit à la vente.')

        if not errors:
            for product, quantity, mode, effective_qty in rows:
                if effective_qty > product.stock:
                    errors.append(
                        f"Stock insuffisant pour {product.name} : "
                        f"demandé {effective_qty}, disponible {product.stock}."
                    )

        # Check credit limit for known customers
        customer = None
        if not errors and customer_phone:
            customer = Customer.objects.filter(phone=customer_phone).first()
        if not errors and not customer and customer_name:
            customer = Customer.objects.filter(name__iexact=customer_name).first()
        if not errors and customer and customer.credit_limit > 0:
            total = sum(product.price * (effective_qty if mode == 'piece' else quantity)
                       for product, quantity, mode, effective_qty in rows)
            if customer.outstanding_balance + total > customer.credit_limit:
                errors.append(
                    f"Limite de crédit dépassée pour {customer.name} : "
                    f"solde {customer.outstanding_balance:.0f} + vente {total:.0f} > "
                    f"plafond {customer.credit_limit:.0f} FCFA"
                )

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            with transaction.atomic():
                sale = Sale.objects.create(
                    sale_date=sale_date, sale_time=timezone.localtime().time(),
                    notes=sale_notes, customer_name=customer_name,
                    customer_phone=customer_phone,
                )
                for product, quantity, mode, effective_qty in rows:
                    if mode in PACK_MODES and product.pack_size > 1:
                        multiplier = PACK_MODES[mode]
                        price = (product.pack_price or (product.price * product.pack_size)) * multiplier
                    else:
                        price = product.price
                    SaleItem.objects.create(
                        sale=sale, product=product,
                        quantity=quantity, price=price,
                        sale_mode=mode
                    )
                    product.stock -= effective_qty
                    product.save(update_fields=['stock'])
                    if product.is_composite and hasattr(product, 'recipe'):
                        for ing in product.recipe.ingredients.all():
                            ing.product.stock -= effective_qty * ing.quantity
                            ing.product.save(update_fields=['stock'])
                sale.update_total()
                sale.amount_paid = sale.total
                sale.sync_payment_status()
                sale.save(update_fields=['amount_paid', 'payment_status'])
            messages.success(request, 'Vente enregistrée avec succès.')
            return redirect('store:sale_detail', sale_id=sale.id)

    return render(request, 'store/create_sale.html', {
        'products': products,
        'quick_products': quick_products,

    })


PACK_MODES = {'paquet': 1, 'carton': 12, 'cartouche': 1}
MODE_LABEL = {'piece': 'pièce', 'paquet': 'paquet', 'carton': 'carton',
              'cartouche': 'cartouche', 'kg': 'kg', 'l': 'litre',
              'sachet': 'sachet', 'boîte': 'boîte'}


def _calc_price(product, qty, mode):
    if mode in PACK_MODES and product.pack_size > 1:
        return (product.pack_price or (product.price * product.pack_size)) * PACK_MODES[mode]
    return product.price


def _calc_effective_qty(product, qty, mode):
    if mode in PACK_MODES and product.pack_size > 1:
        return qty * product.pack_size * PACK_MODES[mode]
    return qty


def _cancel_pending_sale(sale_id):
    """Delete a pending Sale and restore stock for all its items."""
    from django.db import transaction
    try:
        with transaction.atomic():
            sale = Sale.objects.get(id=sale_id)
            for item in sale.items.select_related('product'):
                eq = _calc_effective_qty(item.product, item.quantity, item.sale_mode or 'piece')
                item.product.stock += eq
                item.product.save(update_fields=['stock'])
            sale.delete()
    except Sale.DoesNotExist:
        pass


def _do_quick_sale(product, qty, mode='piece'):
    effective_qty = _calc_effective_qty(product, qty, mode)
    if effective_qty > product.stock:
        return None, f'Stock insuffisant : {product.stock} {product.get_unit_display()}'

    with transaction.atomic():
        sale = Sale.objects.create(
            sale_date=date.today(),
            sale_time=timezone.localtime().time(),
        )
        price = _calc_price(product, qty, mode)
        SaleItem.objects.create(
            sale=sale, product=product,
            quantity=qty, price=price,
            sale_mode=mode
        )
        product.stock -= effective_qty
        product.save(update_fields=['stock'])
        if product.is_composite and hasattr(product, 'recipe'):
            for ing in product.recipe.ingredients.all():
                ing.product.stock -= effective_qty * ing.quantity
                ing.product.save(update_fields=['stock'])
        sale.update_total()
        sale.amount_paid = sale.total
        sale.sync_payment_status()
        sale.save(update_fields=['amount_paid', 'payment_status'])
    return sale, None


def quick_sale(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requis'}, status=405)
    product_id = request.POST.get('product_id')
    quantity = request.POST.get('quantity', '1')
    try:
        product = Product.objects.get(pk=product_id)
        qty = Decimal(quantity)
        if qty <= 0:
            raise InvalidOperation
    except (Product.DoesNotExist, InvalidOperation, ValueError):
        return JsonResponse({'error': 'Produit ou quantité invalide'}, status=400)

    sale, err = _do_quick_sale(product, qty, request.POST.get('mode', 'piece'))
    if err:
        return JsonResponse({'error': err}, status=400)

    return JsonResponse({
        'success': True,
        'product': product.name,
        'quantity': str(qty),
        'total': str(sale.total),
        'sale_id': sale.id,
    })


MODE_KEYWORDS = {
    'piece': ['piece', 'pièce', 'pièces', 'pieces', 'unité', 'unites', 'unités'],
    'paquet': ['paquet', 'paquets', 'paquet'],
    'carton': ['carton', 'cartons'],
    'cartouche': ['cartouche', 'cartouches'],
    'kg': ['kg', 'kilo', 'kilos', 'kilogramme', 'kilogrammes'],
    'l': ['l', 'litre', 'litres'],
}

FRACTIONS = {
    'demi': '0.5', 'demis': '0.5',
    'moitié': '0.5', 'moitie': '0.5', 'moitiés': '0.5',
    'quart': '0.25', 'quarts': '0.25',
    'tiers': '0.333',
}

FRACTION_LABELS = {
    '0.5': 'demi', '0.25': 'quart', '0.333': 'tiers',
}


def _format_qty(qty):
    """Display 0.5 as 'demi', 0.25 as 'quart', etc."""
    s = str(qty)
    if s in FRACTION_LABELS:
        return FRACTION_LABELS[s]
    return s


def _fmt_amount(val):
    """Format Decimal amount nicely: 100.20000 -> '100', 100.50 -> '100.5'"""
    v = Decimal(str(val))
    q = v.quantize(Decimal('0.01'))
    if q == q.to_integral_value():
        return str(int(q))
    return str(q.normalize())


def _action_buttons(actions):
    """Genera HTML con botones clickeables.
    actions = [('label', 'msg'), ...]  →  cada botón envía 'msg' al hacer clic.
    """
    html = '<div class="chat-actions">'
    for label, msg in actions:
        escaped = msg.replace("'", "\\'").replace('"', '&quot;')
        html += f'<button onclick="sendMsg(\'{escaped}\')" class="chat-btn">{label}</button>'
    html += '</div>'
    return html


def _parse_sell_msg(user_msg):
    """Parse a sell command. Returns (product_list, qty, is_price_based, amount, mode) or (None, ...)."""
    sell_keywords = ['vends', 'vendre', 'vend', 'vendez', 'vendons', 'vendu',
                     'achète', 'achete', 'acheter', 'achte', 'achat', 'achats',
                     'prends', 'prend', 'prenez', 'prenons', 'prnd',
                     'donne', 'donner', 'don', 'doner',
                     'vente', 'caisse', 'command', 'commande', 'vds', 'v']
    msg_words = re.split(r'[\s,;\'-]+', user_msg)
    is_sell = any(w in sell_keywords for w in msg_words)
    if not is_sell:
        # Auto-detect sell intent from patterns
        if len(msg_words) <= 3:
            has_nb = bool(re.search(r'\d+[\d,.]*', user_msg))
            has_price = bool(re.search(r'\d+\s*(f(?:cfa)?|francs?|cfa)\b', user_msg, re.I))
            has_mode = bool(re.search(r'\b(pi[eè]ce|pces?|kg|kgs?|l|litre|paquet|carton|sachet|bo[iî]te)\b', user_msg, re.I))
            if has_price or (has_nb and (has_mode or len(msg_words) == 2)):
                is_sell = True
        # Fuzzy typo correction ("end" → "vend", "pnd" → "prends", etc.)
        if not is_sell:
            try:
                corrected = user_msg
                for w in msg_words:
                    match = difflib.get_close_matches(w, sell_keywords, n=1, cutoff=0.55)
                    if match:
                        is_sell = True
                        corrected = re.sub(r'\b' + re.escape(w) + r'\b', match[0], corrected)
                if is_sell:
                    return _parse_sell_msg(corrected)
            except ImportError:
                pass
    if not is_sell:
        return None, None, False, None, None

    mode = None
    qty = Decimal('1')
    rest = user_msg
    for kw in sell_keywords:
        rest = rest.replace(kw, '').strip()
    rest = re.sub(r'^(moi|nous|lui)\s+', '', rest).strip()

    is_price_based = False
    amount = None

    # Extract mode: "2 kg de riz", "3 paquets d'huile", "1 carton"
    for m, words in MODE_KEYWORDS.items():
        for w in words:
            pattern = rf'\b(\d+[\d,.]*)?\s*{re.escape(w)}\b'
            m2 = re.search(pattern, rest, re.IGNORECASE)
            if m2:
                mode = m
                if m2.group(1):
                    try:
                        qty = Decimal(m2.group(1).replace(',', '.'))
                    except:
                        pass
                rest = re.sub(pattern, '', rest, flags=re.IGNORECASE).strip()
                break
        if mode:
            break

    # Extract fractions: "demi pain", "3 quarts de riz"
    for frac, val in FRACTIONS.items():
        pattern = rf'\b(\d+[\d,.]*)?\s*{re.escape(frac)}\b'
        m2 = re.search(pattern, rest, re.IGNORECASE)
        if m2:
            fraction_val = Decimal(val)
            if m2.group(1):
                try:
                    qty = Decimal(m2.group(1).replace(',', '.')) * fraction_val
                except:
                    qty = fraction_val
            else:
                qty = fraction_val
            rest = re.sub(pattern, '', rest, flags=re.IGNORECASE).strip()
            break

    # Try price-based: "huile 100 f" / "500 f de riz" / "100 fcfa" / "100 francs"
    price_match = re.search(r'([\d\s,.]+)\s*(f(?:cfa)?|francs?|cfa)\b', rest, re.IGNORECASE)
    if price_match:
        try:
            amount = Decimal(price_match.group(1).replace(',', '.').replace(' ', ''))
            is_price_based = True
            rest = rest[:price_match.start()] + rest[price_match.end():]
            rest = re.sub(r"\b(de|d')?\s*$", '', rest).strip()
        except:
            pass

    if not is_price_based and mode is None:
        rest = re.sub(r'^un\s+', '1 ', rest).strip()
        rest = re.sub(r'^une\s+', '1 ', rest).strip()
        m = re.match(r'^([\d,.]+)\s+(.+)$', rest)
        if m:
            try:
                qty = Decimal(m.group(1).replace(',', '.'))
                rest = m.group(2).strip()
            except:
                pass
        rest = re.sub(r"^(kg|l|litre|litres|sachet|boite|piece)\s+(de|d')?\s*", '', rest).strip()

    stop_words = {'le', 'la', 'les', 'un', 'une', 'des', 'du', 'de', 'je', 'tu', 'il', 'elle', 'nous', 'vous', 'ils', 'elles', 'ce', 'cet', 'cette', 'ces', 'mon', 'ton', 'son', 'ma', 'ta', 'sa', 'mes', 'tes', 'ses', 'nos', 'vos', 'leur', 'leurs', 'au', 'aux', 'a', 'et', 'ou', 'en', 'dans', 'sur', 'pour', 'par', 'avec', 'sans', 'chez', 'entre', 'moi', 'toi', 'lui', 'eux', 'de', 'd'}
    raw_words = [w for w in re.split(r'[\s,;\'-]+', rest) if w and len(w) > 2 and w not in sell_keywords and w not in stop_words]
    if not raw_words:
        return [], None, is_price_based, amount, mode

    pset = Product.objects.all()
    for w in raw_words:
        variants = {w}
        if w.endswith('s') and len(w) > 3:
            variants.add(w[:-1])
        if w.endswith('x') and len(w) > 3:
            variants.add(w[:-1])
        variant_q = Q()
        for v in variants:
            variant_q |= Q(name__icontains=v) | Q(code__icontains=v) | Q(category__name__icontains=v)
        pset = pset.filter(variant_q)

    if pset.count() == 0:
        # Fallback 1: prefix (startswith) search on each word
        pset = Product.objects.all()
        for w in raw_words:
            pset = pset.filter(Q(name__istartswith=w) | Q(code__istartswith=w))
        if pset.count() > 0:
            return list(pset[:8]), qty, is_price_based, amount, mode
        # Fallback 2: OR match (any word matches, relaxed)
        pset = Product.objects.all()
        or_q = Q()
        for w in raw_words:
            or_q |= Q(name__icontains=w) | Q(code__icontains=w) | Q(category__name__icontains=w)
        pset = pset.filter(or_q)
        if pset.count() > 0:
            return list(pset[:8]), qty, is_price_based, amount, mode
        # Fallback 3: fuzzy matching on product names (full and per-word)
        all_names = list(Product.objects.values_list('name', flat=True))
        fuzzy_set = set()
        # Full text fuzzy match
        full = ' '.join(raw_words)
        fuzzy_set.update(difflib.get_close_matches(full, all_names, n=5, cutoff=0.5))
        # Per-word fuzzy match
        for w in raw_words:
            fuzzy_set.update(difflib.get_close_matches(w, all_names, n=3, cutoff=0.5))
            # Also try matching against each word of product names
            for pn in all_names:
                for pw in pn.lower().split():
                    if difflib.SequenceMatcher(None, w, pw).ratio() >= 0.55:
                        fuzzy_set.add(pn)
        if fuzzy_set:
            fuzzy_ids = list(Product.objects.filter(name__in=list(fuzzy_set)).values_list('id', flat=True)[:8])
            if fuzzy_ids:
                pset = Product.objects.filter(id__in=fuzzy_ids)
                return list(pset[:8]), qty, is_price_based, amount, mode
        return [], None, is_price_based, amount, mode
    return list(pset[:8]), qty, is_price_based, amount, mode


def _add_item_to_existing_sale(sale, product, qty, mode):
    effective_qty = _calc_effective_qty(product, qty, mode)
    if effective_qty > product.stock:
        return f'Stock insuffisant : {product.stock} {product.get_unit_display()}'

    price = _calc_price(product, qty, mode)
    SaleItem.objects.create(
        sale=sale, product=product,
        quantity=qty, price=price,
        sale_mode=mode
    )
    product.stock -= effective_qty
    product.save(update_fields=['stock'])
    if product.is_composite and hasattr(product, 'recipe'):
        for ing in product.recipe.ingredients.all():
            ing.product.stock -= effective_qty * ing.quantity
            ing.product.save(update_fields=['stock'])
    sale.update_total()
    sale.amount_paid = sale.total
    sale.sync_payment_status()
    sale.save(update_fields=['total', 'amount_paid', 'payment_status'])
    return None


def _sale_links(sale_id):
    facture_url = f'/ventes/{sale_id}/facture/'
    pdf_url = f'{facture_url}pdf/'
    ticket_url = f'/ventes/{sale_id}/ticket/'
    return (
        f'<div style="display:flex;gap:10px;margin-top:8px;flex-wrap:wrap;">'
        f'<a href="{facture_url}" target="_blank" style="display:inline-flex;align-items:center;gap:6px;padding:8px 16px;background:#2563eb;color:#fff;border-radius:8px;text-decoration:none;font-size:14px;">📄 Facture</a>'
        f'<a href="{ticket_url}" target="_blank" style="display:inline-flex;align-items:center;gap:6px;padding:8px 16px;background:#16a34a;color:#fff;border-radius:8px;text-decoration:none;font-size:14px;">🧾 Ticket</a>'
        f'<a href="{pdf_url}" target="_blank" style="display:inline-flex;align-items:center;gap:6px;padding:8px 16px;background:#dc2626;color:#fff;border-radius:8px;text-decoration:none;font-size:14px;">📄 PDF</a>'
        f'</div>'
    )


def _suggest_after_sale(product_id):
    """Returns a suggestion string after a sale, or empty string."""
    from store.utils.learning import frequently_bought_together
    together = frequently_bought_together(product_id, limit=2)
    if together:
        parts = ["💡 Aussi demandé :"]
        for t in together:
            icon = '✅' if t['product__stock'] > 0 else '❌'
            parts.append(f"  {icon} {t['product__name']} — {t['product__price']} FCFA")
        return '\n'.join(parts)
    return ''


def chat_assistant(request):
    try:
        return _chat_assistant_impl(request)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'reply': f'❌ Erreur interne : {type(e).__name__}: {str(e)}'})


def _chat_assistant_impl(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requis'}, status=405)
    user_msg = request.POST.get('message', '').strip().lower()
    if not user_msg:
        return JsonResponse({'error': 'Message vide'}, status=400)

    if user_msg in {'efface', 'clear', 'clean', 'nettoie', 'reset'}:
        for k in ['pending_sale', 'pending_cart_sale_id', 'pending_sell_flow', 'pending_stock_adjust',
                  'pending_comma_sale', 'pending_products', 'pending_qty', 'pending_price_based',
                  'pending_amount', 'last_sale_id', 'pending_ask_customer', 'pending_customer_name',
                  'pending_customer_phone', 'pending_ask_mode']:
            request.session.pop(k, None)
        return JsonResponse({'reply': '🧹 Chat et session nettoyés.'})

    cancel_words = {'non', 'no', 'annule', 'cancel', 'oublie', 'rien', 'pas', 'stop', '2'}

    # Clear stale pending_cart_sale_id if it matches a non-existent Sale
    stale_cart = request.session.get('pending_cart_sale_id')
    if stale_cart:
        try:
            Sale.objects.get(id=stale_cart)
        except Sale.DoesNotExist:
            request.session.pop('pending_cart_sale_id', None)

    # --- Guided sell flow (Vendre button or typing "vendre") ---
    sell_trigger = {'vends', 'vendre', 'vend', 'vendez', 'vendons', 'vendu', 'vente', 'vds', 'v'}
    sell_flow = request.session.get('pending_sell_flow')
    if user_msg == '__vendre__' or user_msg in sell_trigger or (sell_flow and user_msg.strip('.,!? ') in cancel_words and not user_msg.strip().isdigit()):
        if sell_flow:
            request.session.pop('pending_sell_flow', None)
        if user_msg in sell_trigger or user_msg == '__vendre__':
            request.session['pending_sell_flow'] = 'product'
            return JsonResponse({'reply': '💰 **Vendre**\nQuel produit veux-tu vendre ?'})
        return JsonResponse({'reply': '❌ Vente annulée.'})

    if sell_flow == 'product':
        # Step 1: user gave product name → search for it
        raw = request.POST.get('message', '').strip()
        # Try spaCy to extract product name from natural language
        try:
            from store.utils.nlp import extract_entities
            prods, _, _, _ = extract_entities(raw)
            search_phrases = [p for p in prods if len(p) > 1]
        except Exception:
            search_phrases = []
        if not search_phrases:
            search_phrases = [w for w in raw.replace(',', ' ').split() if len(w) > 1]
        q = Product.objects.select_related('category')
        for w in search_phrases:
            q = q.filter(Q(name__icontains=w) | Q(code__icontains=w) | Q(category__name__icontains=w))
        results = list(q[:8])
        if not results:
            # Fuzzy fallback
            try:
                all_names = list(Product.objects.values_list('name', flat=True))
                fuzzy_set = set()
                for w in words:
                    close = difflib.get_close_matches(w, all_names, n=3, cutoff=0.5)
                    fuzzy_set.update(close)
                    for pn in all_names:
                        for pw in pn.lower().split():
                            if difflib.SequenceMatcher(None, w, pw).ratio() >= 0.55:
                                fuzzy_set.add(pn)
                if fuzzy_set:
                    results = list(Product.objects.filter(name__in=list(fuzzy_set))[:8])
            except ImportError:
                pass
        if not results:
            request.session['pending_sell_flow'] = 'product'
            return JsonResponse({'reply': f'❌ Produit introuvable. Réessaie un autre nom ou tape "annule" pour quitter.'})
        if len(results) == 1:
            p = results[0]
            request.session['pending_sell_flow'] = {'step': 'qty', 'product_id': p.id}
            mode_btns = _action_buttons([['1 Pièce', '1 piece'], ['1 Kg', '1 kg'], ['1 Litre', '1 litre'], ['1 Paquet', '1 paquet'], ['1 Carton', '1 carton']])
            return JsonResponse({'reply': f'💰 **{p.name}** — {p.price} FCFA (stock: {p.stock})\nCombien ? (ex: **2**, **500f**, **2 kg**, **demi litre**){mode_btns}'})
        # Multiple results
        names = '\n'.join(f"{i+1}. {r.name} — {r.price} FCFA (stock: {r.stock})" for i, r in enumerate(results))
        request.session['pending_sell_flow'] = {'step': 'pick', 'ids': [r.id for r in results]}
        btns = _action_buttons([(f'{r.name[:20]}', str(i+1)) for i, r in enumerate(results)])
        return JsonResponse({'reply': f'Plusieurs produits correspondent. Lequel ?\n{names}\n{btns}'})

    if isinstance(sell_flow, dict):
        sf = sell_flow
        if sf.get('step') == 'pick':
            if user_msg.isdigit():
                n = int(user_msg)
                ids = sf.get('ids', [])
                if 1 <= n <= len(ids):
                    p = Product.objects.get(id=ids[n-1])
                    request.session['pending_sell_flow'] = {'step': 'qty', 'product_id': p.id}
                    mode_btns = _action_buttons([['1 Pièce', '1 piece'], ['1 Kg', '1 kg'], ['1 Litre', '1 litre'], ['1 Paquet', '1 paquet'], ['1 Carton', '1 carton']])
                    return JsonResponse({'reply': f'💰 **{p.name}** — {p.price} FCFA (stock: {p.stock})\nCombien ? (ex: **2**, **500f**, **2 kg**, **demi litre**){mode_btns}'})
                return JsonResponse({'reply': f'Numéro invalide.'})
            # User typed a name instead of number → search among listed products
            ids = sf.get('ids', [])
            matched = [Product.objects.get(id=pid) for pid in ids if pid and Product.objects.filter(id=pid, name__icontains=user_msg).exists()]
            if len(matched) == 1:
                p = matched[0]
                request.session['pending_sell_flow'] = {'step': 'qty', 'product_id': p.id}
                mode_btns = _action_buttons([['1 Pièce', '1 piece'], ['1 Kg', '1 kg'], ['1 Litre', '1 litre'], ['1 Paquet', '1 paquet'], ['1 Carton', '1 carton']])
                return JsonResponse({'reply': f'💰 **{p.name}** — {p.price} FCFA (stock: {p.stock})\nCombien ? (ex: **2**, **500f**, **2 kg**, **demi litre**){mode_btns}'})
            prods_list = Product.objects.filter(id__in=ids)
            names = '\n'.join(f"{i+1}. {p.name} — {p.price} FCFA (stock: {p.stock})" for i, p in enumerate(prods_list))
            btns = _action_buttons([(f'{p.name[:20]}', str(i+1)) for i, p in enumerate(prods_list)])
            return JsonResponse({'reply': f'❌ "{user_msg}" introuvable. Choisis dans la liste :\n{names}\n{btns}'})
        elif sf.get('step') == 'qty':
            # Step 2: user gave quantity → confirm and sell
            p = Product.objects.get(id=sf['product_id'])
            request.session.pop('pending_sell_flow', None)
            # Parse qty/mode from user message
            has_price = bool(re.search(r'\d+[\d,.]*\s*(f(?:cfa)?|francs?|cfa)\b', user_msg, re.I))
            if has_price:
                # Price-based: extract amount
                m = re.search(r'(\d+[\d,.]*)\s*(f(?:cfa)?|francs?|cfa)', user_msg, re.I)
                if m:
                    amount = Decimal(m.group(1).replace(',', '.'))
                    if p.price <= 0:
                        return JsonResponse({'reply': f'❌ **{p.name}** a un prix de 0 FCFA. Impossible de vendre par montant.'})
                    qty = amount / p.price
                    mode = None
                    # Check for mode word in message
                    mode_map = {'piece': 'piece', 'kg': 'kg', 'kilo': 'kg', 'litre': 'l', 'l': 'l', 'paquet': 'paquet', 'carton': 'carton'}
                    for kw, mv in mode_map.items():
                        if kw in user_msg:
                            mode = mv
                            break
                    effective = _calc_effective_qty(p, qty, mode or 'piece')
                    if effective > p.stock:
                        request.session['pending_stock_adjust'] = {'product_id': p.id, 'qty': str(qty), 'mode': mode or 'piece'}
                        return JsonResponse({'reply': f'⚠️ Stock insuffisant : {p.stock} {p.get_unit_display()} dispo, besoin de ~{effective:.2f} (soit {_format_qty(qty)} {mode or "pièce"}).\nQue faire ?' + _action_buttons([['📦 Ajouter du stock', '1'], ['✅ Vendre le disponible', '2'], ['❌ Annuler', '3']])})
                    request.session['pending_sale'] = {'product_id': p.id, 'qty': str(qty), 'mode': mode or 'piece'}
                    total = _calc_price(p, qty, mode or 'piece') * qty
                    
                    ml = MODE_LABEL.get(mode, 'pièce') if mode else 'pièce'
                    return JsonResponse({'reply': f'Vendre ~{_format_qty(qty)} {ml} de **{p.name}** à {p.price} FCFA = {_fmt_amount(qty * p.price)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})
            # Quantity-based: extract number and mode
            nums = re.findall(r'(\d+[\d,.]*)', user_msg)
            qty = Decimal(nums[0].replace(',', '.')) if nums else Decimal('1')
            mode = None
            user_provided_mode = False
            frac_map = {'demi': '0.5', 'quart': '0.25', 'tiers': '0.333', 'moitié': '0.5', 'moitie': '0.5'}
            for kw, fv in frac_map.items():
                if kw in user_msg:
                    qty = Decimal(fv)
                    break
            mode_map = {'piece': 'piece', 'kg': 'kg', 'kilo': 'kg', 'litre': 'l', 'l': 'l', 'paquet': 'paquet', 'carton': 'carton'}
            for kw, mv in mode_map.items():
                if kw in user_msg:
                    mode = mv
                    user_provided_mode = True
                    break
            if mode is None:
                unit_map = {'kg': 'kg', 'l': 'l', 'pièce': 'piece', 'sachet': 'piece', 'boîte': 'piece'}
                mode = unit_map.get(p.unit, 'piece')
            effective = _calc_effective_qty(p, qty, mode or 'piece')
            # If bare number (no mode, no f/cfa) exceeds stock, try as FCFA amount
            if not user_provided_mode and effective > p.stock and p.price > 0:
                amount_try = qty
                qty_as_amount = amount_try / p.price
                eff_as_amount = _calc_effective_qty(p, qty_as_amount, mode or 'piece')
                if eff_as_amount <= p.stock:
                    qty = qty_as_amount
                    effective = eff_as_amount
                    is_price_based = True
            if effective > p.stock:
                request.session['pending_stock_adjust'] = {'product_id': p.id, 'qty': str(qty), 'mode': mode or 'piece'}
                return JsonResponse({'reply': f'⚠️ Stock insuffisant : {p.stock} {p.get_unit_display()} dispo, besoin de ~{effective:.2f} (soit {_format_qty(qty)} {mode or "pièce"}).\nQue faire ?' + _action_buttons([['📦 Ajouter du stock', '1'], ['✅ Vendre le disponible', '2'], ['❌ Annuler', '3']])})
            total = _calc_price(p, qty, mode) * qty
            request.session['pending_sale'] = {'product_id': p.id, 'qty': str(qty), 'mode': mode}
            
            return JsonResponse({'reply': f'Vendre **{_format_qty(qty)}** {MODE_LABEL.get(mode, "pièce")} de **{p.name}** = {_fmt_amount(total)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})
        # Fallback: clear
        request.session.pop('pending_sell_flow', None)

    # --- Number selection (multi-product / print choice) ---
    if 'pending_stock_adjust' in request.session and not user_msg.isdigit():
        # Non-digit message while stock adjust is pending → cancel
        request.session.pop('pending_stock_adjust', None)
    if user_msg.isdigit() and 'pending_customer_name' not in request.session and 'pending_customer_phone' not in request.session:
        n = int(user_msg)
        # Pending stock adjustment
        if 'pending_stock_adjust' in request.session:
            pending = request.session.pop('pending_stock_adjust')
            product = Product.objects.get(id=pending['product_id'])
            qty = Decimal(pending['qty'])
            mode = pending.get('mode', 'piece')
            if n == 1:
                supply_url = f"/produits/{product.id}/approvisionner/"
                return JsonResponse({'reply': f'📦 Va sur la page d\'approvisionnement : {supply_url}\nAjoute le stock, puis tape "vends {product.name}" pour continuer.'})
            elif n == 2:
                sell_qty = min(qty, Decimal(str(product.stock)))
                if mode:
                    sale, err = _do_quick_sale(product, sell_qty, mode)
                else:
                    sale, err = _do_quick_sale(product, sell_qty, 'piece')
                if err:
                    return JsonResponse({'reply': f'❌ {err}'})
                request.session['last_sale_id'] = sale.id
                return JsonResponse({'reply': f'✅ **{product.name}** — {sell_qty} vendu(s) sur {qty} demandé(s)\n💰 Total : {_fmt_amount(sale.total)} FCFA\n{_sale_links(sale.id)}'})
            else:
                return JsonResponse({'reply': '❌ Vente annulée.'})
        # Pending comma-separated multi-item sale confirmation
        if 'pending_comma_sale' in request.session:
            if n == 1:
                pending = request.session.pop('pending_comma_sale')
                try:
                    with transaction.atomic():
                        sale = Sale.objects.create(sale_date=date.today(), sale_time=timezone.localtime().time())
                        lines = []
                        for item in pending:
                            p = Product.objects.get(id=item['product_id'])
                            iq = Decimal(item['qty'])
                            im = item.get('mode', 'piece')
                            pr = _calc_price(p, iq, im)
                            eq = _calc_effective_qty(p, iq, im)
                            SaleItem.objects.create(sale=sale, product=p, quantity=iq, price=pr, sale_mode=im)
                            p.stock -= eq
                            p.save(update_fields=['stock'])
                            subtotal = (iq * pr).quantize(Decimal('0.01'))
                            lines.append(f"• **{p.name}** — {iq} {MODE_LABEL.get(im, im)} = {_fmt_amount(subtotal)} FCFA")
                        sale.update_total()
                        sale.amount_paid = sale.total
                        sale.sync_payment_status()
                        sale.save(update_fields=['total', 'amount_paid', 'payment_status'])
                    request.session['pending_cart_sale_id'] = sale.id
                    last_p = Product.objects.get(id=pending[-1]['product_id'])
                    suggest = _suggest_after_sale(last_p.id)
                    reply = f'✅ **Vente #{sale.id}** — {len(pending)} article{"s" if len(pending) > 1 else ""}\n' + '\n'.join(lines)
                    reply += f'\n💰 **Total : {_fmt_amount(sale.total)} FCFA**'
                    if suggest:
                        reply += f'\n{suggest}'
                    reply += '\n\n➕ Ajouter un autre article ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])
                    return JsonResponse({'reply': reply})
                except Exception as e:
                    return JsonResponse({'reply': f'❌ Erreur : {str(e)}'})
            elif n == 2:
                request.session.pop('pending_comma_sale', None)
                return JsonResponse({'reply': 'Vente annulée.'})
        # Pending products selection
        if 'pending_products' in request.session:
            prods = request.session.pop('pending_products')
            if 1 <= n <= len(prods):
                product = Product.objects.get(id=prods[n - 1])
                qty = Decimal(request.session.pop('pending_qty', '1'))
                is_price_based = request.session.pop('pending_price_based', False)
                amount = request.session.pop('pending_amount', None)
                if is_price_based and amount:
                    amount = Decimal(amount)
                    if product.price <= 0:
                        return JsonResponse({'reply': f'❌ **{product.name}** a un prix de 0 FCFA. Impossible de vendre par montant.'})
                    qty = amount / product.price
                    if qty > product.stock:
                        request.session['pending_stock_adjust'] = {'product_id': product.id, 'qty': str(qty), 'mode': 'piece'}
                        supply_url = f"/produits/{product.id}/approvisionner/"
                        return JsonResponse({'reply': f'⚠️ Stock insuffisant : {product.stock} {product.get_unit_display()} dispo, besoin de ~{qty:.2f}.\nQue faire ?' + _action_buttons([['📦 Ajouter du stock', '1'], ['✅ Vendre le disponible', '2'], ['❌ Annuler', '3']])})
                    qty = qty.quantize(Decimal('0.001'))
                total = qty * product.price
                request.session.pop('pending_cart_sale_id', None)
                request.session['pending_sale'] = {'product_id': product.id, 'qty': str(qty)}
                return JsonResponse({'reply': f'Vendre **{_format_qty(qty)}** {product.get_unit_display()} de **{product.name}** = {_fmt_amount(total)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})
            else:
                return JsonResponse({'reply': f'Numéro invalide. Choisis entre 1 et {len(prods)}.'})
        # Print choice after sale
        if 'last_sale_id' in request.session:
            sale_id = request.session.pop('last_sale_id')
            if n == 1:
                return JsonResponse({'reply': f'{_sale_links(sale_id)}\n\n📄 Facture ouverte. Tape "ticket" pour le ticket de caisse.'})
            elif n == 2:
                return JsonResponse({'reply': f'{_sale_links(sale_id)}\n\n🧾 Ticket ouvert. Tape "facture" pour la facture.'})
            elif n >= 3:
                return JsonResponse({'reply': '✅ Vente enregistrée.'})
        # "Ajouter ?" response by number (1=oui, 2=non)
        if 'pending_cart_sale_id' in request.session and 'pending_sale' not in request.session and 'pending_products' not in request.session and 'last_sale_id' not in request.session and 'pending_ask_customer' not in request.session:
            if n == 1:
                request.session['pending_sell_flow'] = 'product'
                return JsonResponse({'reply': '👍 Quel produit veux-tu ajouter ?'})
            elif n == 2:
                sale_id = request.session.pop('pending_cart_sale_id', None)
                if sale_id:
                    _cancel_pending_sale(sale_id)
                return JsonResponse({'reply': '✅ Vente annulée.'})
        # Confirmation by number
        if 'pending_sale' in request.session:
            if n == 1:
                try:
                    pending = request.session.pop('pending_sale')
                    product = Product.objects.get(id=pending['product_id'])
                    qty = Decimal(pending['qty'])
                    sale_mode = pending.get('mode', 'piece')
                    cart_sale_id = request.session.get('pending_cart_sale_id')
                    if cart_sale_id:
                        try:
                            sale = Sale.objects.get(id=cart_sale_id)
                        except Sale.DoesNotExist:
                            request.session.pop('pending_cart_sale_id', None)
                            cart_sale_id = None
                    if cart_sale_id:
                        err = _add_item_to_existing_sale(sale, product, qty, sale_mode)
                        if err and 'Stock insuffisant' in err:
                            request.session['pending_stock_adjust'] = {'product_id': product.id, 'qty': str(qty), 'mode': sale_mode}
                            return JsonResponse({'reply': f'⚠️ {err}\nQue faire ?' + _action_buttons([['📦 Ajouter du stock', '1'], ['✅ Vendre le disponible', '2'], ['❌ Annuler', '3']])})
                        elif err:
                            return JsonResponse({'reply': f'❌ {err}'})
                    else:
                        sale, err = _do_quick_sale(product, qty, sale_mode)
                        if err and 'Stock insuffisant' in err:
                            request.session['pending_stock_adjust'] = {'product_id': product.id, 'qty': str(qty), 'mode': sale_mode}
                            return JsonResponse({'reply': f'⚠️ {err}\nQue faire ?' + _action_buttons([['📦 Ajouter du stock', '1'], ['✅ Vendre le disponible', '2'], ['❌ Annuler', '3']])})
                        elif err:
                            return JsonResponse({'reply': f'❌ {err}'})
                        request.session['pending_cart_sale_id'] = sale.id
                    suggest = _suggest_after_sale(product.id)
                    total_now = sale.total if cart_sale_id else (product.price * qty)
                    reply = f'✅ **{product.name}** ({qty} {MODE_LABEL.get(sale_mode, sale_mode)})'
                    reply += f'\n💰 **Total vente : {_fmt_amount(total_now)} FCFA** ({sale.items.count()} article{"s" if sale.items.count() > 1 else ""})'
                    if suggest:
                        reply += f'\n{suggest}'
                    reply += f'\n{_sale_links(sale.id)}'
                    reply += '\n\n➕ Ajouter un autre article ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])
                    return JsonResponse({'reply': reply})
                except Exception as e:
                    return JsonResponse({'reply': f'❌ Erreur lors de la vente : {str(e)}'})
            elif n == 2:
                request.session.pop('pending_sale', None)
                return JsonResponse({'reply': 'Vente annulée.'})

    # --- Print/text request for last sale ---
    if 'last_sale_id' in request.session:
        sale_id = request.session['last_sale_id']
        clean = user_msg.strip('.,!? ')
        if clean in {'facture', 'ticket', 'recu', 'imprime', 'imprimer', 'pdf'} or any(w in clean for w in ['facture', 'ticket']):
            request.session.pop('last_sale_id', None)
            return JsonResponse({'reply': f'{_sale_links(sale_id)}'})
        else:
            request.session.pop('last_sale_id', None)

    # --- Customer info response ---
    if 'pending_ask_customer' in request.session:
        sale_id = request.session.pop('pending_ask_customer')
        request.session['pending_cart_sale_id'] = sale_id
        if user_msg.isdigit():
            n = int(user_msg)
            if n == 2:
                request.session.pop('pending_cart_sale_id')
                request.session['last_sale_id'] = sale_id
                return JsonResponse({'reply': f'✅ Vente sans client.\n{_sale_links(sale_id)}'})
        clean = user_msg.strip('.,!? ')
        if clean in {'non', 'no', 'skip', 'pas', 'stop', '2'}:
            request.session.pop('pending_cart_sale_id', None)
            request.session['last_sale_id'] = sale_id
            return JsonResponse({'reply': f'✅ Vente sans client.\n{_sale_links(sale_id)}'})
        else:
            request.session['pending_customer_name'] = sale_id
            return JsonResponse({'reply': '👤 **Infos client**\nQuel est le nom du client ? (tape "non" pour ignorer)'})

    # --- Confirmation / cancellation ---
    clean_msg = user_msg.strip('.,!? ')
    confirm_words = {'oui', 'yes', 'y', 'ok', 'confirme', 'valide', 'vas-y', 'vazy', 'bien', 'daccord', '1'}

    # Handle "non" after "Ajouter un autre article ?" — cancel entire sale
    if clean_msg in cancel_words and 'pending_cart_sale_id' in request.session and 'pending_sale' not in request.session:
        sale_id = request.session.pop('pending_cart_sale_id', None)
        if sale_id:
            _cancel_pending_sale(sale_id)
        return JsonResponse({'reply': '✅ Vente annulée.'})

    # Handle "oui" after "Ajouter un autre article ?"
    if clean_msg in confirm_words and 'pending_cart_sale_id' in request.session and 'pending_sale' not in request.session:
        request.session['pending_sell_flow'] = 'product'
        return JsonResponse({'reply': '👍 Quel produit veux-tu ajouter ?'})

    if clean_msg in confirm_words or clean_msg == 'y':
        pending = request.session.pop('pending_sale', None)
        if pending:
            try:
                product = Product.objects.get(id=pending['product_id'])
                qty = Decimal(pending['qty'])
                sale_mode = pending.get('mode', 'piece')
                cart_sale_id = request.session.get('pending_cart_sale_id')
                if cart_sale_id:
                    try:
                        sale = Sale.objects.get(id=cart_sale_id)
                    except Sale.DoesNotExist:
                        request.session.pop('pending_cart_sale_id', None)
                        cart_sale_id = None
                if cart_sale_id:
                    err = _add_item_to_existing_sale(sale, product, qty, sale_mode)
                    if err:
                        return JsonResponse({'reply': f'❌ {err}'})
                else:
                    sale, err = _do_quick_sale(product, qty, sale_mode)
                    if err:
                        return JsonResponse({'reply': f'❌ {err}'})
                    request.session['pending_cart_sale_id'] = sale.id
                suggest = _suggest_after_sale(product.id)
                total_now = sale.total if cart_sale_id else (product.price * qty)
                n_items = sale.items.count()
                reply = f'✅ **{product.name}** ({qty} {MODE_LABEL.get(sale_mode, sale_mode)})'
                reply += f'\n💰 **Total vente : {_fmt_amount(total_now)} FCFA** ({n_items} article{"s" if n_items > 1 else ""})'
                if suggest:
                    reply += f'\n{suggest}'
                reply += '\n\n➕ Ajouter un autre article ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])
                return JsonResponse({'reply': reply})
            except Product.DoesNotExist:
                return JsonResponse({'reply': '❌ Produit introuvable (supprimé ?).'})
            except Exception as e:
                return JsonResponse({'reply': f'❌ Erreur lors de la vente : {str(e)}'})
        return JsonResponse({'reply': 'Aucune vente en attente.'})

    if clean_msg in cancel_words:
        if 'pending_sale' in request.session:
            request.session.pop('pending_sale', None)
            return JsonResponse({'reply': 'Vente annulée.'})

    # --- Learning-based intents ---
    from store.utils.learning import (top_selling, trending, time_based_suggestions,
                                      low_stock_velocity_alerts, suggestions_for_cart,
                                      frequently_bought_together)

    if any(w in user_msg for w in ['suggère', 'suggere', 'recommande', 'propose', 'conseil', 'idée', 'idee', 'quoi vendre']):
        hour = timezone.localtime().hour
        period = 'matin' if hour < 12 else 'après-midi' if hour < 18 else 'soir'
        now_suggestions = time_based_suggestions(limit=4)
        top = top_selling(period_days=7, limit=4)

        reply = f"🤖 **Suggestions pour ce {period}**\n\n"
        if now_suggestions:
            reply += f"⏰ Produits qui se vendent en ce moment :\n"
            for s in now_suggestions:
                reply += f"  • {s['product__name']} — {s['product__price']} FCFA\n"
        reply += f"\n🏆 Meilleures ventes de la semaine :\n"
        for s in top:
            reply += f"  • {s['product__name']} — {s['total_qty']} vendus\n"
        return JsonResponse({'reply': reply})

    if any(w in user_msg for w in ['populaire', 'meilleure vente', 'meilleur', 'top', 'plus vendu']):
        top = top_selling(period_days=30, limit=8)
        reply = "🏆 **Top ventes (30 jours)**\n"
        for i, s in enumerate(top, 1):
            reply += f"{i}. {s['product__name']} — {s['total_qty']} vendus — {s['total_revenue']} FCFA\n"
        return JsonResponse({'reply': reply})

    if any(w in user_msg for w in ['tendance', 'trend', 'qui monte', 'explose']):
        trend = trending(period_days=7, limit=5)
        reply = "📈 **Tendances de la semaine**\n"
        for t in trend:
            arrow = '🟢' if t['growth_pct'] > 0 else '🔴'
            if t['past_qty'] == 0 and t['recent_qty'] > 0:
                arrow = '🆕'
            reply += f"{arrow} {t['name']} : {t['recent_qty']} vendus"
            if t['past_qty'] > 0:
                reply += f" (+{t['growth_pct']}% vs semaine précédente)"
            reply += "\n"
        if not trend:
            reply += "Pas assez de données pour détecter des tendances."
        return JsonResponse({'reply': reply})

    if any(w in user_msg for w in ['alerte', 'reappro', 'réappro', 'rupture', 'urgent', 'bientot fini']):
        alerts = low_stock_velocity_alerts(threshold_days=7, limit=5)
        reply = "⚠️ **Alertes stocks**\n"
        if alerts:
            for a in alerts:
                reply += f"• {a['name']} : {a['stock']} restants — ≈ {a['days_left']} jours avant rupture\n"
        else:
            reply += "Aucun produit en risque de rupture cette semaine."
        return JsonResponse({'reply': reply})

    if any(w in user_msg for w in ['habitude', 'apprentissage', 'que tu sais', 'appris']):
        top = top_selling(period_days=7, limit=3)
        trend = trending(period_days=7, limit=2)
        reply = "🧠 **Ce que j'ai appris de tes habitudes**\n\n"
        reply += "Je me base sur toutes tes ventes passées pour :\n"
        reply += "• Suggérer des produits qui vont bien ensemble\n"
        reply += "• Recommander ce qui se vend le mieux\n"
        reply += "• Détecter les tendances émergentes\n"
        reply += "• T'alerter avant les ruptures de stock\n\n"
        if top:
            reply += f"Cette semaine tes clients achètent surtout : {', '.join(t['product__name'] for t in top)}\n"
        if trend:
            reply += f"Produit qui monte : {trend[0]['name']} (+{trend[0]['growth_pct']}%)"
        return JsonResponse({'reply': reply})

    # --- Nouvel article, approvisionnement ---
    if any(w in user_msg for w in ['ajouter article', 'nouveau produit', 'créer', 'creer', 'ajout article', 'nouvel article']):
        return JsonResponse({'reply': f'📦 **Ajouter un article** : <a href="/produits/gestion/nouveau/" target="_blank">Ouvrir le formulaire</a>\nTu peux aussi taper le nom du produit pour voir son stock.'})

    if any(w in user_msg for w in ['approvisionner', 'alimenter stock', 'réappro', 'reappro', 'stock appro']):
        return JsonResponse({'reply': f'📦 **Approvisionner un produit** :\nTape le nom du produit que tu veux approvisionner (ex: "riz") ou va sur <a href="/produits/gestion/" target="_blank">la gestion des produits</a>.'})

    # --- Customer info collection after sale ---
    if 'pending_customer_name' in request.session:
        if clean_msg in cancel_words:
            request.session.pop('pending_customer_name', None)
            sale_id = request.session.pop('pending_cart_sale_id', None)
            if sale_id:
                request.session['last_sale_id'] = sale_id
                return JsonResponse({'reply': f'✅ Vente sans client.\n{_sale_links(sale_id)}'})
            return JsonResponse({'reply': '✅ Vente enregistrée.'})
        customer_name = user_msg.strip().title()
        if len(customer_name) < 2:
            return JsonResponse({'reply': 'Nom trop court. Quel est le nom du client ?'})
        request.session['pending_customer_phone'] = customer_name
        request.session.pop('pending_customer_name', None)
        return JsonResponse({'reply': f'👤 Client : **{customer_name}**\n📞 Son téléphone ? (ou tape "non" pour ignorer)'})

    if 'pending_customer_phone' in request.session:
        customer_name = request.session.pop('pending_customer_phone')
        phone = '' if clean_msg in cancel_words else user_msg.strip()
        sale_id = request.session.pop('pending_cart_sale_id', None)
        if sale_id:
            try:
                sale = Sale.objects.get(id=sale_id)
                sale.customer_name = customer_name
                sale.customer_phone = phone
                sale.save(update_fields=['customer_name', 'customer_phone'])
            except Sale.DoesNotExist:
                pass
        request.session['last_sale_id'] = sale_id
        reply = f'✅ Client **{customer_name}** enregistré.'
        if sale_id:
            reply += f'\n{_sale_links(sale_id)}'
        return JsonResponse({'reply': reply})

    # --- Analyse de texte (prix, produits) ---
    _price_pattern = re.findall(r'(\w[\w\s]*?)\s+(\d+[\d,.]*)\s*(f(?:cfa)?|francs?|cfa)\s*(?:[/le]+\s*)?(\w+)?', user_msg, re.I)
    if len(_price_pattern) >= 1 and not any(w in user_msg for w in ['vends', 'vendre', 'achète', 'acheter', 'prix', 'coûte', 'combien']):
        parts = []
        for match in _price_pattern[:5]:
            name = match[0].strip().lower()
            price = match[1].replace(',', '.')
            unit = match[3] if match[3] else ''
            prod = Product.objects.filter(name__icontains=name).first()
            if prod:
                parts.append(f"• **{prod.name}** — {price} FCFA/{unit} (stock: {prod.stock})")
            else:
                parts.append(f"• {name.title()} — {price} FCFA/{unit} (nouveau produit)")
        if parts:
            reply = "📋 **Prix relevés :**\n" + '\n'.join(parts) + '\n\nTape "vends [nom]" pour vendre ou "ajouter article" pour créer un nouveau produit.'
            return JsonResponse({'reply': reply})

    words = [w for w in user_msg.replace(',', ' ').split() if len(w) > 1]

    # --- Handle follow-up for mode selection ---
    if 'pending_ask_mode' in request.session:
        pending = request.session.pop('pending_ask_mode')
        product = Product.objects.get(id=pending['product_id'])
        qty = Decimal(pending['qty'])
        user_mode = None
        # Handle number selection: "1" = Pièce, "2" = Kg, "3" = Litre, "4" = Paquet, "5" = Carton
        if user_msg.strip().isdigit():
            n = int(user_msg.strip())
            mode_map = {1: 'piece', 2: 'kg', 3: 'l', 4: 'paquet', 5: 'carton'}
            if n in mode_map:
                user_mode = mode_map[n]
        # Try to extract qty and mode from the user's answer
        import re as _re
        mode_qty_match = _re.match(r'^(\d+[\d,.]*)\s*(.+)$', user_msg.strip())
        if mode_qty_match:
            try:
                qty = Decimal(mode_qty_match.group(1).replace(',', '.'))
                mode_text = mode_qty_match.group(2).strip()
                for m, words in MODE_KEYWORDS.items():
                    if any(w in mode_text for w in words):
                        user_mode = m
                        break
            except:
                pass
        # Also check fraction words in the answer
        for frac, val in FRACTIONS.items():
            if frac in user_msg:
                qty = Decimal(val)
                break
        if not user_mode:
            for m, words in MODE_KEYWORDS.items():
                if any(w in user_msg for w in words):
                    user_mode = m
                    break
        if not user_mode:
            user_mode = 'piece'
        total = qty * product.price
        request.session.pop('pending_cart_sale_id', None)
        request.session['pending_sale'] = {
            'product_id': product.id,
            'qty': str(qty),
            'mode': user_mode
        }
        return JsonResponse({'reply': f'Vendre **{_format_qty(qty)}** {MODE_LABEL.get(user_mode,user_mode)} de **{product.name}** = {_fmt_amount(total)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})

    # --- Multi-item quick sale (comma-separated) ---
    sell_keywords = ['vends', 'vendre', 'vend', 'vendez', 'vendons', 'vendu',
                     'achète', 'achete', 'acheter', 'achte', 'achat', 'achats',
                     'prends', 'prend', 'prenez', 'prenons', 'prnd',
                     'donne', 'donner', 'don', 'doner',
                     'vente', 'caisse', 'command', 'commande', 'vds', 'v']
    msg_words = re.split(r'[\s,;\'-]+', user_msg)
    if ',' in user_msg and any(w in user_msg for w in sell_keywords):
        raw_items = [i.strip() for i in user_msg.split(',') if i.strip()]
        if len(raw_items) >= 2:
            parsed = []
            for item in raw_items:
                sub = f"vends {item}" if not any(w in item for w in sell_keywords) else item
                prods, iq, ipb, iamt, im = _parse_sell_msg(sub)
                if prods is None or not prods:
                    return JsonResponse({'reply': f'❌ "{item}" : produit introuvable.'})
                if len(prods) > 1:
                    return JsonResponse({'reply': f'❌ "{item}" correspond à plusieurs produits.'})
                p = prods[0]
                if ipb:
                    if iamt <= Decimal('0'):
                        continue
                    iq = (iamt / p.price).quantize(Decimal('0.001'))
                iq = iq or Decimal('1')
                im = im or 'piece'
                if _calc_effective_qty(p, iq, im) > p.stock:
                    return JsonResponse({'reply': f'⚠️ Stock insuffisant pour **{p.name}** : {p.stock} {p.get_unit_display()} dispo.'})
                parsed.append((p, iq, im, _calc_price(p, iq, im)))
            if not parsed:
                return JsonResponse({'reply': '❌ Aucun article valide.'})
            # Store in session for confirmation, don't create Sale yet
            pending = []
            for p, iq, im, pr in parsed:
                pending.append({'product_id': p.id, 'qty': str(iq), 'mode': im})
            request.session['pending_comma_sale'] = pending
            lines = []
            for p, iq, im, pr in parsed:
                subtotal = (iq * pr).quantize(Decimal('0.01'))
                lines.append(f"• **{p.name}** — {iq} {MODE_LABEL.get(im, im)} = {_fmt_amount(subtotal)} FCFA")
            total = sum(iq * pr for p, iq, im, pr in parsed)
            reply = f'Vendre ces {len(parsed)} articles ?\n' + '\n'.join(lines)
            reply += f'\n💰 **Total : {_fmt_amount(total)} FCFA**'
            reply += '\n' + _action_buttons([['✅ Confirmer', '1'], ['❌ Annuler', '2']])
            return JsonResponse({'reply': reply})

    # --- Detect & parse sell command ---
    products, qty, is_price_based, amount, mode = _parse_sell_msg(user_msg)
    if products is not None:
        if len(products) == 0:
            # Extract name from user message for new product suggestion
            raw_name = ' '.join(w for w in msg_words if w not in sell_keywords and len(w) > 2)
            if raw_name:
                add_url = f"/produits/ajouter/?name={raw_name}"
                return JsonResponse({'reply': f'❌ Produit introuvable : "{raw_name}".\nQue faire ?' + _action_buttons([['➕ Ajouter ce produit', f'ajouter article {raw_name}'], ['❌ Annuler', 'annule']])})
            return JsonResponse({'reply': '❌ Produit introuvable. Tape "ajouter article" pour en créer un nouveau.'})
        if len(products) == 1:
            product = products[0]
        else:
            request.session['pending_products'] = [p.id for p in products[:8]]
            request.session['pending_qty'] = str(qty)
            request.session['pending_price_based'] = is_price_based
            request.session['pending_amount'] = str(amount) if amount else None
            names = '\n'.join(f"{i+1}. {p.name} — {p.price} FCFA (stock: {p.stock})" for i, p in enumerate(products[:8]))
            btns = _action_buttons([(f'{i+1}. {p.name[:20]}', str(i+1)) for i, p in enumerate(products[:8])])
            return JsonResponse({'reply': f'Plusieurs produits correspondent. Lequel ?\n{names}\n{btns}'})

        if is_price_based:
            if amount <= 0:
                return JsonResponse({'reply': '❌ Montant invalide.'})
            if product.price <= 0:
                return JsonResponse({'reply': f'❌ **{product.name}** a un prix de 0 FCFA. Impossible de vendre par montant.'})
            qty = amount / product.price
            if qty > product.stock:
                request.session['pending_stock_adjust'] = {'product_id': product.id, 'qty': str(qty), 'mode': mode or 'piece'}
                return JsonResponse({'reply': f'⚠️ Stock insuffisant : {product.stock} {product.get_unit_display()} dispo, besoin de ~{qty:.2f}.\nQue faire ?' + _action_buttons([['📦 Ajouter du stock', '1'], ['✅ Vendre le disponible', '2'], ['❌ Annuler', '3']])})
            qty = qty.quantize(Decimal('0.001'))
            m = mode or 'piece'
            request.session.pop('pending_cart_sale_id', None)
            request.session['pending_sale'] = {'product_id': product.id, 'qty': str(qty), 'mode': m}
            return JsonResponse({'reply': f'Vendre ~{_format_qty(qty)} {MODE_LABEL.get(m,m)} de **{product.name}** à {product.price} FCFA/{MODE_LABEL.get(m,m)} = {_fmt_amount(amount)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})
        else:
            if mode:
                from decimal import ROUND_HALF_UP
                effective_qty = qty
                if product.pack_size > 1:
                    effective_qty = (qty * product.pack_size).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                total = qty * product.price
                request.session.pop('pending_cart_sale_id', None)
                request.session['pending_sale'] = {'product_id': product.id, 'qty': str(qty), 'mode': mode}
                return JsonResponse({'reply': f'Vendre **{_format_qty(qty)}** {MODE_LABEL.get(mode,mode)} de **{product.name}** = {_fmt_amount(total)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})
            else:
                unit_map = {'kg': 'kg', 'l': 'l', 'pièce': 'piece', 'sachet': 'piece', 'boîte': 'piece'}
                inferred = unit_map.get(product.unit, 'piece')
                if qty and qty > 0:
                    request.session.pop('pending_cart_sale_id', None)
                    request.session['pending_sale'] = {'product_id': product.id, 'qty': str(qty), 'mode': inferred}
                    total = qty * product.price
                    return JsonResponse({'reply': f'Vendre **{_format_qty(qty)}** {MODE_LABEL.get(inferred,inferred)} de **{product.name}** = {_fmt_amount(total)} FCFA ?' + _action_buttons([['✅ Oui', '1'], ['❌ Non', '2']])})
                request.session['pending_ask_mode'] = {
                    'product_id': product.id,
                    'qty': str(qty or '1'),
                }
                mode_btns = _action_buttons([['Pièce', 'piece 1'], ['Kg', 'kg 1'], ['Litre', 'litre 1'], ['Paquet', 'paquet 1'], ['Carton', 'carton 1']])
                return JsonResponse({'reply': f'Pour **{product.name}**, combien et en quoi ?{mode_btns}\n'
                                              f'Réponds par exemple : "2 kg", "3 pièces", "1 litre", "demi", "quart"…'})

    # --- Normal search ---
    q = Product.objects.select_related('category')
    for w in words:
        if w in sell_keywords:
            continue
        q = q.filter(
            Q(name__icontains=w) | Q(code__icontains=w) |
            Q(category__name__icontains=w)
        )
    results = q[:10]

    is_stock = any(w in user_msg for w in ['stock', 'combien', 'reste', 'disponible', 'quantité', 'rupture'])
    is_price = any(w in user_msg for w in ['prix', 'coûte', 'tarif', 'combien', 'frais', 'paye', 'vaut'])
    is_category = any(w in user_msg for w in ['catégorie', 'rayon', 'famille', 'type', 'liste', 'tous', 'tout'])
    is_low = any(w in user_msg for w in ['faible', 'épuisé', 'alerte', 'critique', 'minimum', 'réapprovisionner'])

    reply_parts = []

    if is_low and q.count() == 0:
        low_products = Product.objects.filter(stock__lte=F('min_stock'))
        if low_products:
            reply_parts.append('⚠️ Produits avec stock faible ou épuisé :')
            for p in low_products[:10]:
                reply_parts.append(f"- {p.name} : {p.stock} {p.get_unit_display()}")
        else:
            reply_parts.append('Aucun produit en stock critique.')
        return JsonResponse({'reply': '\n'.join(reply_parts)})

    if results:
        if is_category or (len(results) > 5 and not is_stock and not is_price):
            cats = {}
            for p in results:
                cn = p.category.name if p.category else 'Sans catégorie'
                cats.setdefault(cn, []).append(p.name)
            reply_parts.append(f"Produits trouvés ({len(results)}) :")
            for cat, names in cats.items():
                reply_parts.append(f"\n📦 {cat} :")
                for n in names[:5]:
                    reply_parts.append(f"  • {n}")
        else:
            for p in results:
                line = f"• {p.name}"
                if is_price or not is_stock:
                    line += f" — {p.price} FCFA"
                if is_stock or not is_price:
                    line += f" — stock: {p.stock} {p.get_unit_display()}"
                if is_low and p.min_stock:
                    line += f" (min: {p.min_stock})"
                reply_parts.append(line)
            if len(results) == 10:
                reply_parts.append(f"\net d'autres produits correspondent...")
            # Suggest complementary products if only one product found
            if len(results) == 1 and not is_stock and not is_low:
                together = frequently_bought_together(results[0].id, limit=3)
                if together:
                    reply_parts.append(f"\n💡 Souvent achetés avec **{results[0].name}** :")
                    for t in together:
                        icon = '✅' if t['product__stock'] > 0 else '❌'
                        reply_parts.append(f"  {icon} {t['product__name']} — {t['product__price']} FCFA")
    else:
        reply_parts.append("Aucun produit trouvé. Essaie d'autres mots-clés.")
        cats = Category.objects.all()
        if cats:
            reply_parts.append(f"\nCatégories disponibles : {', '.join(c.name for c in cats)}")
        # Suggest top selling when search yields nothing
        top = top_selling(period_days=7, limit=4)
        if top:
            reply_parts.append(f"\n💡 Meilleures ventes de la semaine :")
            for s in top:
                reply_parts.append(f"  • {s['product__name']} — {s['total_qty']} vendus")

    return JsonResponse({'reply': '\n'.join(reply_parts)})


def sale_history(request):
    sort = request.GET.get('sort', '')
    order = request.GET.get('order', 'desc')

    sales = Sale.objects.prefetch_related('items')
    sort_fields = {
        'id': 'id',
        'date': 'sale_date',
        'total': 'total',
        'items': 'id',
    }
    if sort in sort_fields:
        order_prefix = '-' if order == 'desc' else ''
        sales = sales.order_by(f'{order_prefix}{sort_fields[sort]}')
    else:
        sales = sales.order_by('-created_at')

    today = date.today()
    daily_total = Sale.objects.filter(sale_date=today).aggregate(total=Sum('total'))['total'] or 0
    monthly_total = Sale.objects.filter(sale_date__year=today.year, sale_date__month=today.month).aggregate(total=Sum('total'))['total'] or 0
    total_amount = Sale.objects.aggregate(total=Sum('total'))['total'] or 0

    paginator = Paginator(sales, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']

    all_sales_count = Sale.objects.count()
    report_text = (
        f"Rapport des ventes - {today.strftime('%d/%m/%Y')}\n"
        f"Ventes du jour: {daily_total} FCFA\n"
        f"Ventes du mois: {monthly_total} FCFA\n"
        f"Total general: {total_amount} FCFA\n"
        f"Nombre de ventes: {all_sales_count}"
    )
    return render(request, 'store/sale_history.html', {
        'page_obj': page_obj,
        'daily_total': daily_total,
        'monthly_total': monthly_total,
        'total_amount': total_amount,
        'whatsapp_report_url': f"https://wa.me/?text={quote(report_text)}",
        'email_report_url': f"mailto:?subject={quote('Rapport des ventes')}&body={quote(report_text)}",
        'current_sort': sort,
        'current_order': order,
        'query_params': qp.urlencode(),
    })


def _restore_stock_and_delete_sales(sales):
    with transaction.atomic():
        for sale in sales.prefetch_related('items__product'):
            for item in sale.items.all():
                qty = item.quantity
                if item.sale_mode in ('paquet', 'carton', 'cartouche') and item.product.pack_size > 1:
                    mult = 12 if item.sale_mode == 'carton' else 1
                    qty = item.quantity * item.product.pack_size * mult
                item.product.stock += qty
                item.product.save(update_fields=['stock'])
            sale.delete()


def sale_detail(request, sale_id):
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    composite_items = []
    for item in sale.items.all():
        if item.product.is_composite and hasattr(item.product, 'recipe'):
            breakdown = []
            for ing in item.product.recipe.ingredients.all():
                breakdown.append({
                    'name': ing.product.name,
                    'qty_per_unit': ing.quantity,
                    'total_consumed': item.quantity * ing.quantity,
                })
            composite_items.append({
                'item': item,
                'breakdown': breakdown,
            })
    return render(request, 'store/sale_detail.html', {
        'sale': sale,
        'composite_items': composite_items,
    })


def sale_update(request, sale_id):
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    products = Product.objects.select_related('category').all()

    original_quantities = {}
    for item in sale.items.all():
        original_qtys = Decimal('0')
        if item.sale_mode in ('paquet', 'carton', 'cartouche') and item.product.pack_size > 1:
            mult = 12 if item.sale_mode == 'carton' else 1
            original_qtys = item.quantity * item.product.pack_size * mult
        else:
            original_qtys = item.quantity
        original_quantities[item.product_id] = original_quantities.get(item.product_id, Decimal('0')) + original_qtys

    if request.method == 'POST':
        product_ids = request.POST.getlist('product')
        quantities = request.POST.getlist('quantity')
        sale_modes = request.POST.getlist('sale_mode')
        sale_date_str = request.POST.get('sale_date', '')
        sale_notes = request.POST.get('notes', '')
        customer_name = request.POST.get('customer_name', '').strip()
        customer_phone = request.POST.get('customer_phone', '').strip()
        rows = []
        errors = []

        try:
            sale.sale_date = date.fromisoformat(sale_date_str) if sale_date_str else date.today()
        except (ValueError, TypeError):
            sale.sale_date = date.today()
        sale.notes = sale_notes
        sale.customer_name = customer_name
        sale.customer_phone = customer_phone

        for index, product_id in enumerate(product_ids):
            quantity_value = quantities[index] if index < len(quantities) else ''
            mode = sale_modes[index] if index < len(sale_modes) else 'piece'
            if not product_id and not quantity_value:
                continue
            try:
                product = Product.objects.get(pk=product_id)
                quantity = Decimal(quantity_value)
                if quantity <= 0:
                    raise InvalidOperation
            except (Product.DoesNotExist, InvalidOperation, ValueError):
                errors.append('Une ligne de vente est invalide.')
                continue

            if mode in ('paquet', 'carton', 'cartouche') and product.pack_size > 1:
                multiplier = 12 if mode == 'carton' else 1
                effective_qty = quantity * product.pack_size * multiplier
            else:
                effective_qty = quantity
                if mode not in ('kg', 'l'):
                    mode = 'piece'

            rows.append((product, quantity, mode, effective_qty))

        if not rows:
            errors.append('Ajoute au moins un produit à la vente.')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            with transaction.atomic():
                for item in sale.items.select_related('product'):
                    if item.sale_mode in ('paquet', 'carton', 'cartouche') and item.product.pack_size > 1:
                        mult = 12 if item.sale_mode == 'carton' else 1
                        item.product.stock += item.quantity * item.product.pack_size * mult
                    else:
                        item.product.stock += item.quantity
                    item.product.save(update_fields=['stock'])

                sale.items.all().delete()

                for product, quantity, mode, effective_qty in rows:
                    if mode in ('paquet', 'carton', 'cartouche') and product.pack_size > 1:
                        multiplier = 12 if mode == 'carton' else 1
                        price = (product.pack_price or (product.price * product.pack_size)) * multiplier
                    else:
                        price = product.price
                    SaleItem.objects.create(
                        sale=sale, product=product,
                        quantity=quantity, price=price,
                        sale_mode=mode
                    )
                    product.stock -= effective_qty
                    product.save(update_fields=['stock'])

                sale.save(update_fields=['sale_date', 'notes'])
                sale.update_total()

            messages.success(request, 'Vente modifiée avec succès.')
            return redirect('store:sale_detail', sale_id=sale.id)

    sale_rows = [
        {'product_id': item.product_id, 'quantity': item.quantity, 'sale_mode': item.sale_mode}
        for item in sale.items.all()
    ]
    return render(request, 'store/sale_form.html', {
        'sale': sale,
        'products': products,
        'sale_rows': sale_rows,
        'title': f'Modifier la vente #{sale.id}',
        'submit_label': 'Enregistrer les modifications',
    })


def sale_delete(request, sale_id):
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    if request.method == 'POST':
        _restore_stock_and_delete_sales(Sale.objects.filter(pk=sale.pk))
        messages.success(request, 'Vente supprimée et stock restauré.')
        return redirect('store:sale_history')
    return render(request, 'store/sale_confirm_delete.html', {'sale': sale})


def sale_bulk_delete(request):
    if request.method != 'POST':
        return redirect('store:sale_history')

    action = request.POST.get('action')
    selected_ids = request.POST.getlist('sale_ids')

    if action == 'delete_all':
        sales = Sale.objects.all()
        count = sales.count()
        _restore_stock_and_delete_sales(sales)
        messages.success(request, f'{count} vente(s) supprimée(s). Stock restauré.')
    elif action == 'delete_selected':
        sales = Sale.objects.filter(id__in=selected_ids)
        count = sales.count()
        if count:
            _restore_stock_and_delete_sales(sales)
            messages.success(request, f'{count} vente(s) sélectionnée(s) supprimée(s). Stock restauré.')
        else:
            messages.warning(request, 'Sélectionne au moins une vente à supprimer.')
    else:
        messages.warning(request, 'Action de suppression invalide.')

    return redirect('store:sale_history')


def sales_report_pdf(request):
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    today = date.today()
    sales = Sale.objects.prefetch_related('items__product').order_by('-created_at')
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 45
    content_width = width - 2 * margin
    primary = rl_colors.HexColor('#7f1d1d')
    y = height - 50

    pdf.setFillColor(primary)
    pdf.roundRect(margin - 5, y - 5, content_width + 10, 55, 8, fill=1, stroke=0)
    pdf.setFillColor(rl_colors.white)
    pdf.setFont('Helvetica-Bold', 20)
    pdf.drawString(margin + 10, y + 18, 'Rapport des ventes')
    pdf.setFont('Helvetica', 10)
    pdf.drawString(margin + 10, y + 4, today.strftime('Généré le %d/%m/%Y'))
    y -= 70

    total_global = 0
    total_count = sales.count()
    for sale in sales:
        total_global += float(sale.total)

    pdf.setFillColor(rl_colors.HexColor('#334155'))
    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(margin, y, f'Résumé')
    y -= 22
    pdf.setFont('Helvetica', 11)
    pdf.drawString(margin, y, f'Total des ventes: {int(total_global):,} FCFA'.replace(',', ' '))
    y -= 16
    pdf.drawString(margin, y, f'Nombre de ventes: {total_count}')
    y -= 28

    pdf.setStrokeColor(rl_colors.HexColor('#e2e8f0'))
    pdf.line(margin, y, margin + content_width, y)
    y -= 20

    pdf.setFont('Helvetica-Bold', 14)
    pdf.setFillColor(primary)
    pdf.drawString(margin, y, 'Détail des ventes')
    y -= 25

    pdf.setFont('Helvetica-Bold', 9)
    pdf.setFillColor(rl_colors.HexColor('#64748b'))
    cols = [('N°', 30), ('Client', 120), ('Total', 80), ('Payé', 80), ('Statut', 80), ('Méthode', 80), ('Date', 100)]
    x_pos = margin
    for label, w in cols:
        pdf.drawString(x_pos, y, label)
        x_pos += w
    y -= 14

    for sale in sales:
        if y < 60:
            pdf.showPage()
            y = height - 50
        pdf.setFont('Helvetica', 9)
        pdf.setFillColor(rl_colors.HexColor('#334155'))
        x_pos = margin
        vals = [
            str(sale.id),
            sale.customer_name or (sale.customer.name if sale.customer else '—'),
            f'{int(float(sale.total)):,} FCFA'.replace(',', ' '),
            f'{int(float(sale.amount_paid)):,} FCFA'.replace(',', ' '),
            sale.payment_status or '—',
            sale.payment_method or '—',
            sale.created_at.strftime('%d/%m/%Y %H:%M') if sale.created_at else '—',
        ]
        for val, (_, w) in zip(vals, cols):
            pdf.drawString(x_pos, y, val[:w//5])
            x_pos += w
        y -= 12

    pdf.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'rapport-ventes-{today.isoformat()}.pdf')


def sale_invoice(request, sale_id):
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    return render(request, 'store/sale_invoice.html', {'sale': sale, 'settings': StoreSettings.load()})


def sale_receipt(request, sale_id):
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    currency_symbols = {'XOF': 'FCFA', 'EUR': '€', 'USD': '$'}
    return render(request, 'store/sale_receipt.html', {
        'sale': sale,
        'settings': StoreSettings.load(),
        'currency_symbol': currency_symbols.get(sale.currency, 'FCFA'),
    })


def sale_invoice_pdf(request, sale_id):
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from .utils import montant_en_lettres

    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    settings = StoreSettings.load()
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 45
    content_width = width - 2 * margin

    # Couleurs
    primary = rl_colors.HexColor('#2563eb')
    primary_dark = rl_colors.HexColor('#1d4ed8')
    gray_50 = rl_colors.HexColor('#f8fafc')
    gray_200 = rl_colors.HexColor('#e2e8f0')
    gray_500 = rl_colors.HexColor('#64748b')
    gray_700 = rl_colors.HexColor('#334155')

    y = height - 40

    # En-tête (bandeau coloré)
    pdf.setFillColor(primary)
    pdf.setStrokeColor(primary)
    pdf.setLineWidth(0)
    pdf.roundRect(margin - 5, y - 5, content_width + 10, 72, 8, fill=1, stroke=0)
    pdf.setFillColor(rl_colors.white)

    # Logo + nom magasin
    logo_x = margin + 8
    if settings.logo:
        try:
            pdf.drawImage(settings.logo.path, logo_x, y + 8, width=70, height=35, preserveAspectRatio=True, mask='auto')
            logo_x += 80
        except Exception:
            pass

    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(logo_x, y + 22, settings.store_name)
    pdf.setFont('Helvetica', 8)
    store_info = settings.address or ''
    if settings.phone_number:
        store_info += f"  Tel: {settings.phone_number}" if store_info else f"Tel: {settings.phone_number}"
    if store_info:
        pdf.drawString(logo_x, y + 8, store_info)

    # N° facture à droite
    pdf.setFont('Helvetica-Bold', 22)
    pdf.drawRightString(width - margin - 8, y + 22, f'#{sale.id}')
    pdf.setFont('Helvetica', 9)
    pdf.drawRightString(width - margin - 8, y + 6, f'Date: {sale.sale_date.strftime("%d/%m/%Y")}')

    y -= 60

    # Client
    if sale.customer_name:
        pdf.setFillColor(gray_500)
        pdf.setFont('Helvetica', 7)
        pdf.drawString(margin, y, 'CLIENT')
        pdf.setFillColor(gray_700)
        pdf.setFont('Helvetica-Bold', 11)
        y -= 14
        pdf.drawString(margin, y, sale.customer_name)
        y -= 14
        if sale.customer_phone:
            pdf.setFont('Helvetica', 9)
            pdf.setFillColor(gray_500)
            pdf.drawString(margin, y, f'Tel: {sale.customer_phone}')
            y -= 18
        else:
            y -= 6

    if sale.notes:
        y -= 8
        pdf.setFillColor(rl_colors.HexColor('#9a3412'))
        pdf.setFont('Helvetica-Oblique', 9)
        pdf.drawString(margin, y, f'Note: {sale.notes}')
        y -= 18

    # Tableau en-tête
    y -= 8
    pdf.setFillColor(primary)
    pdf.roundRect(margin, y - 4, content_width, 18, 4, fill=1, stroke=0)
    pdf.setFillColor(rl_colors.white)
    pdf.setFont('Helvetica-Bold', 8)
    col_x = [margin + 8, margin + 70, margin + 270, margin + 340, margin + 390, margin + 460]
    headers = ['Code', 'Produit', 'Prix', 'Qté', 'Mode', 'Total']
    for i, h in enumerate(headers):
        pdf.drawString(col_x[i], y, h)
    y -= 26

    # Lignes du tableau
    pdf.setFont('Helvetica', 9)
    for item in sale.items.all():
        if y < 70:
            pdf.showPage()
            y = height - 40
        pdf.setFillColor(gray_700)
        pdf.drawString(col_x[0], y, item.product.code or '-')
        pdf.drawString(col_x[1], y, item.product.name[:28])
        pdf.drawRightString(col_x[2] + 50, y, f'{item.price:,.0f} F')
        pdf.drawString(col_x[3], y, f'{item.quantity:,.0f} {item.product.get_unit_display()}')
        mode_text = 'Pqt' if item.sale_mode == 'paquet' else 'Dét'
        pdf.drawString(col_x[4], y, mode_text)
        pdf.drawRightString(col_x[5] + 50, y, f'{item.subtotal:,.0f} F')
        y -= 18

    # Ligne de total
    y -= 6
    pdf.setStrokeColor(primary)
    pdf.setLineWidth(1.5)
    pdf.line(margin, y, width - margin, y)
    y -= 22
    pdf.setFillColor(primary)
    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawRightString(width - margin, y, f'Total: {sale.total:,.0f} FCFA')

    # Montant en lettres
    y -= 26
    pdf.setFillColor(gray_500)
    pdf.setFont('Helvetica-Oblique', 9)
    montant_lettres = montant_en_lettres(sale.total)
    pdf.drawString(margin, y, f'Arrêtée la présente facture à la somme de : {montant_lettres}')

    # Signature
    y -= 40
    if settings.signature:
        try:
            pdf.drawImage(settings.signature.path, margin, y - 25, width=100, height=40, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    else:
        pdf.setFont('Helvetica', 9)
        pdf.setFillColor(gray_500)
        pdf.drawString(margin, y, 'Cachet & signature')

    pdf.setFont('Helvetica', 7)
    pdf.setFillColor(gray_500)
    pdf.drawRightString(width - margin, y, f'Facture #{sale.id} - {sale.sale_date.strftime("%d/%m/%Y")}')

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="facture_vente_{sale.id}.pdf"'
    return response


def export_sales_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="ventes.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Date', 'Total'])
    for sale in Sale.objects.order_by('-created_at'):
        writer.writerow([sale.id, timezone.localtime(sale.created_at).strftime('%d/%m/%Y %H:%M'), sale.total])
    return response


def _get_cart(request):
    return request.session.setdefault('cart', {})


def add_to_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    form = CartAddForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        quantity = form.cleaned_data['quantity']
        cart = _get_cart(request)
        key = str(product.id)
        cart[key] = str(Decimal(cart.get(key, '0')) + quantity)
        request.session.modified = True
        messages.success(request, 'Produit ajouté au panier.')
    return redirect('store:product_list')


def remove_from_cart(request, product_id):
    cart = _get_cart(request)
    cart.pop(str(product_id), None)
    request.session.modified = True
    messages.success(request, 'Produit retiré du panier.')
    return redirect('store:cart')


def cart_view(request):
    cart = _get_cart(request)
    products = Product.objects.filter(id__in=cart.keys())
    items = []
    total = Decimal('0')

    for product in products:
        quantity = Decimal(cart.get(str(product.id), '0'))
        subtotal = product.price * quantity
        items.append({'product': product, 'quantity': quantity, 'subtotal': subtotal})
        total += subtotal

    return render(request, 'store/cart.html', {'items': items, 'total': total})


def cart_checkout(request):
    cart = _get_cart(request)
    if not cart:
        messages.warning(request, 'Le panier est vide.')
        return redirect('store:cart')

    products = {str(product.id): product for product in Product.objects.filter(id__in=cart.keys())}

    for product_id, quantity in cart.items():
        product = products[product_id]
        quantity = Decimal(quantity)
        if quantity > product.stock:
            messages.error(request, f"Stock insuffisant pour {product.name} : demandé {quantity}, disponible {product.stock}.")
            return redirect('store:cart')

    with transaction.atomic():
        sale = Sale.objects.create()
        for product_id, quantity in cart.items():
            product = products[product_id]
            quantity = Decimal(quantity)
            SaleItem.objects.create(sale=sale, product=product, quantity=quantity, price=product.price)
            product.stock -= quantity
            product.save(update_fields=['stock'])
        sale.update_total()
        sale.amount_paid = sale.total
        sale.sync_payment_status()
        sale.save(update_fields=['amount_paid', 'payment_status'])

    request.session['cart'] = {}
    messages.success(request, 'Panier validé avec succès.')
    return redirect('store:sale_detail', sale_id=sale.id)


def expense_list(request):
    sort = request.GET.get('sort', '')
    order = request.GET.get('order', 'desc')
    expenses = Expense.objects.all()
    sort_fields = {
        'description': 'description',
        'amount': 'amount',
        'category': 'category',
        'date': 'date',
    }
    if sort in sort_fields:
        order_prefix = '-' if order == 'desc' else ''
        expenses = expenses.order_by(f'{order_prefix}{sort_fields[sort]}')
    else:
        expenses = expenses.order_by('-date')
    total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    settings = StoreSettings.load()
    today = date.today()

    current_month_expenses = Expense.objects.filter(date__year=today.year, date__month=today.month)
    current_month_total = current_month_expenses.aggregate(total=Sum('amount'))['total'] or 0
    monthly_limit = settings.monthly_expense_limit or 0
    over_budget = monthly_limit > 0 and current_month_total > monthly_limit
    budget_remaining = monthly_limit - current_month_total if monthly_limit else 0

    category_rows = current_month_expenses.values('category').annotate(total=Sum('amount')).order_by('-total')
    category_labels = []
    category_values = []
    category_display = dict(Expense.CATEGORY_CHOICES)
    for row in category_rows:
        category_labels.append(category_display.get(row['category'], row['category'] or 'Autre'))
        category_values.append(float(row['total'] or 0))

    monthly_totals = {m: 0 for m in range(1, 13)}
    rows = Expense.objects.filter(date__year=today.year).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('amount'))
    for row in rows:
        monthly_totals[row['month'].month] = float(row['total'] or 0)

    paginator = Paginator(expenses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']

    return render(request, 'store/expense_list.html', {
        'page_obj': page_obj,
        'total_expenses': total_expenses,
        'current_month_total': current_month_total,
        'monthly_limit': monthly_limit,
        'budget_remaining': budget_remaining,
        'over_budget': over_budget,
        'category_labels': json.dumps(category_labels),
        'category_values': json.dumps(category_values),
        'month_labels': json.dumps([date(1900, m, 1).strftime('%b') for m in range(1, 13)]),
        'month_values': json.dumps([monthly_totals[m] for m in range(1, 13)]),
        'current_sort': sort,
        'current_order': order,
        'query_params': qp.urlencode(),
    })


def expense_create(request):
    form = ExpenseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Dépense ajoutée avec succès.')
        return redirect('store:expense_list')
    return render(request, 'store/expense_form.html', {'form': form, 'title': 'Nouvelle dépense'})


def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    form = ExpenseForm(request.POST or None, instance=expense)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Dépense modifiée avec succès.')
        return redirect('store:expense_list')
    return render(request, 'store/expense_form.html', {'form': form, 'title': 'Modifier la dépense'})


def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Dépense supprimée avec succès.')
        return redirect('store:expense_list')
    return render(request, 'store/expense_confirm_delete.html', {'expense': expense})


def debt_list(request):
    sort = request.GET.get('sort', '')
    order = request.GET.get('order', 'asc')
    debts = Debt.objects.all()
    sort_fields = {
        'person': 'person',
        'amount': 'amount',
        'due_date': 'due_date',
        'type': 'debt_type',
        'status': 'paid',
    }
    if sort in sort_fields:
        order_prefix = '-' if order == 'desc' else ''
        debts = debts.order_by(f'{order_prefix}{sort_fields[sort]}')
    else:
        debts = debts.order_by('paid', 'due_date')

    outstanding_debt_total = Debt.objects.filter(paid=False).aggregate(total=Sum('amount'))['total'] or 0
    overdue_debt_total = Debt.objects.filter(paid=False, due_date__lt=date.today()).aggregate(total=Sum('amount'))['total'] or 0

    paginator = Paginator(debts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']

    return render(request, 'store/debt_list.html', {
        'page_obj': page_obj,
        'today': date.today(),
        'outstanding_debt_total': outstanding_debt_total,
        'overdue_debt_total': overdue_debt_total,
        'current_sort': sort,
        'current_order': order,
        'query_params': qp.urlencode(),
    })


def debt_create(request):
    form = DebtForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Dette ajoutée avec succès.')
        return redirect('store:debt_list')
    return render(request, 'store/debt_form.html', {'form': form, 'title': 'Nouvelle dette'})


def debt_update(request, pk):
    debt = get_object_or_404(Debt, pk=pk)
    form = DebtForm(request.POST or None, instance=debt)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Dette modifiée avec succès.')
        return redirect('store:debt_list')
    return render(request, 'store/debt_form.html', {'form': form, 'title': 'Modifier la dette'})


def debt_delete(request, pk):
    debt = get_object_or_404(Debt, pk=pk)
    if request.method == 'POST':
        debt.delete()
        messages.success(request, 'Dette supprimée avec succès.')
        return redirect('store:debt_list')
    return render(request, 'store/debt_confirm_delete.html', {'debt': debt})


def debt_mark_paid(request, pk):
    debt = get_object_or_404(Debt, pk=pk)
    debt.paid = True
    debt.paid_at = timezone.now()
    debt.save(update_fields=['paid', 'paid_at'])
    messages.success(request, 'Dette marquée comme réglée.')
    return redirect('store:debt_list')


def phone_credit_list(request):
    phone_credits = PhoneCredit.objects.all()
    total_phone_credits = PhoneCredit.objects.aggregate(total=Sum('amount'))['total'] or 0
    paginator = Paginator(phone_credits, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']
    return render(request, 'store/phone_credit_list.html', {
        'page_obj': page_obj,
        'total_phone_credits': total_phone_credits,
        'query_params': qp.urlencode(),
    })


def phone_credit_create(request):
    form = PhoneCreditForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Crédit téléphonique ajouté avec succès.')
        return redirect('store:phone_credit_list')
    return render(request, 'store/phone_credit_form.html', {'form': form, 'title': 'Nouveau crédit'})


def phone_credit_update(request, pk):
    phone_credit = get_object_or_404(PhoneCredit, pk=pk)
    form = PhoneCreditForm(request.POST or None, instance=phone_credit)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Crédit téléphonique modifié avec succès.')
        return redirect('store:phone_credit_list')
    return render(request, 'store/phone_credit_form.html', {'form': form, 'title': 'Modifier le crédit'})


def phone_credit_delete(request, pk):
    phone_credit = get_object_or_404(PhoneCredit, pk=pk)
    if request.method == 'POST':
        phone_credit.delete()
        messages.success(request, 'Crédit téléphonique supprimé avec succès.')
        return redirect('store:phone_credit_list')
    return render(request, 'store/phone_credit_confirm_delete.html', {'phone_credit': phone_credit})


def phone_credit_purchase_list(request):
    purchases = PhoneCreditPurchase.objects.all()
    total_purchased = PhoneCreditPurchase.objects.aggregate(total=Sum('amount'))['total'] or 0
    available_stock = PhoneCreditPurchase.get_available_stock()
    paginator = Paginator(purchases, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']
    return render(request, 'store/phone_credit_purchase_list.html', {
        'page_obj': page_obj,
        'total_purchased': total_purchased,
        'available_stock': available_stock,
        'low_stock_alert': available_stock < Decimal('10000.00'),
        'query_params': qp.urlencode(),
    })


def phone_credit_purchase_create(request):
    form = PhoneCreditPurchaseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Achat de crédit ajouté avec succès.')
        return redirect('store:phone_credit_purchase_list')
    return render(request, 'store/phone_credit_purchase_form.html', {'form': form, 'title': 'Nouvel achat'})


def phone_credit_purchase_update(request, pk):
    purchase = get_object_or_404(PhoneCreditPurchase, pk=pk)
    form = PhoneCreditPurchaseForm(request.POST or None, instance=purchase)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Achat de crédit modifié avec succès.')
        return redirect('store:phone_credit_purchase_list')
    return render(request, 'store/phone_credit_purchase_form.html', {'form': form, 'title': 'Modifier l’achat'})


def phone_credit_purchase_delete(request, pk):
    purchase = get_object_or_404(PhoneCreditPurchase, pk=pk)
    if request.method == 'POST':
        purchase.delete()
        messages.success(request, 'Achat de crédit supprimé avec succès.')
        return redirect('store:phone_credit_purchase_list')
    return render(request, 'store/phone_credit_purchase_confirm_delete.html', {'purchase': purchase})


# =========================
# PERTES DE STOCK
# =========================
def stock_loss_list(request):
    losses = StockLoss.objects.select_related('product').all()
    total_loss = StockLoss.objects.aggregate(total=Sum('loss_amount'))['total'] or 0
    paginator = Paginator(losses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    qp = request.GET.copy()
    if 'page' in qp:
        del qp['page']
    return render(request, 'store/stock_loss_list.html', {
        'page_obj': page_obj,
        'total_loss': total_loss,
        'query_params': qp.urlencode(),
    })


def stock_loss_create(request):
    form = StockLossForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        loss = form.save()
        product = loss.product
        if product.stock >= loss.quantity:
            product.stock -= loss.quantity
            product.save()
        messages.success(request, 'Perte enregistrée avec succès.')
        return redirect('store:stock_loss_list')
    return render(request, 'store/stock_loss_form.html', {
        'form': form,
        'title': 'Nouvelle perte',
    })


def stock_loss_expired(request):
    today = date.today()
    expired_products = Product.objects.filter(
        expiry_date__lt=today,
        stock__gt=0
    ).order_by('expiry_date')

    if request.method == 'POST':
        product_ids = request.POST.getlist('products')
        for pid in product_ids:
            product = get_object_or_404(Product, pk=pid)
            if product.stock > 0:
                StockLoss.objects.create(
                    product=product,
                    quantity=product.stock,
                    loss_amount=product.stock * (product.cost_price or 0),
                    reason='expired',
                    notes=f'Périmé le {product.expiry_date}'
                )
                product.stock = 0
                product.save()
        messages.success(request, f'{len(product_ids)} produit(s) périmé(s) marqué(s) comme perte.')
        return redirect('store:stock_loss_list')

    return render(request, 'store/stock_loss_expired.html', {
        'expired_products': expired_products,
        'today': today,
    })


# =========================
# BÉNÉFICES
# =========================
def profit_view(request):
    start_date = request.GET.get('start_date', str(date.today().replace(month=1, day=1)))
    end_date = request.GET.get('end_date', str(date.today()))
    items = SaleItem.objects.filter(sale__sale_date__gte=start_date, sale__sale_date__lte=end_date).select_related('product')
    per_product = {}
    total_revenue = Decimal('0')
    total_cost = Decimal('0')

    PACK_MODES = {'paquet': 1, 'carton': 12, 'cartouche': 1}

    for item in items:
        revenue = item.subtotal
        if item.sale_mode in PACK_MODES and item.product.pack_size > 1:
            mult = PACK_MODES[item.sale_mode]
            effective_qty = item.quantity * item.product.pack_size * mult
        else:
            effective_qty = item.quantity
        cost = effective_qty * (item.product.cost_price or 0)
        profit = revenue - cost

        total_revenue += revenue
        total_cost += cost

        pid = item.product_id
        if pid not in per_product:
            per_product[pid] = {
                'product': item.product,
                'qty': Decimal('0'),
                'revenue': Decimal('0'),
                'cost': Decimal('0'),
                'profit': Decimal('0'),
            }
        per_product[pid]['qty'] += effective_qty
        per_product[pid]['revenue'] += revenue
        per_product[pid]['cost'] += cost
        per_product[pid]['profit'] += profit

    for p in per_product.values():
        p['margin'] = (p['profit'] / p['revenue'] * 100) if p['revenue'] else 0

    sorted_products = sorted(per_product.values(), key=lambda x: -x['profit'])
    paginator = Paginator(sorted_products, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_profit = total_revenue - total_cost
    margin = (total_profit / total_revenue * 100) if total_revenue else 0

    return render(request, 'store/profit_list.html', {
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'margin': margin,
    })


# ===== PRIX EN LOT =====
def bulk_price_update(request):
    if request.method == 'POST':
        raw_ids = request.POST.get('product_ids', '')
        new_price_raw = request.POST.get('new_price', '')
        price_type = request.POST.get('price_type', 'price')
        try:
            if raw_ids == 'all':
                products = Product.objects.all()
            else:
                ids = [int(x) for x in raw_ids.split(',') if x.strip().isdigit()]
                products = Product.objects.filter(id__in=ids) if ids else Product.objects.none()

            # Support percentage: "+10%" or "-15%"
            new_price_raw = new_price_raw.strip()
            if new_price_raw.endswith('%'):
                pct = Decimal(new_price_raw.rstrip('%'))
                count = 0
                for p in products:
                    current = getattr(p, price_type) or Decimal('0')
                    new_val = current * (Decimal('1') + pct / Decimal('100'))
                    setattr(p, price_type, new_val)
                    p.save(update_fields=[price_type])
                    count += 1
            else:
                new_price = Decimal(new_price_raw)
                count = products.update(**{price_type: new_price})
            messages.success(request, f'{count} produit(s) mis à jour.')
        except (ValueError, TypeError, InvalidOperation):
            messages.error(request, 'Valeur invalide.')
        return redirect('store:product_manage')
    return redirect('store:product_manage')


# ===== POS TACTILE =====
def pos_view(request):
    from .models import Promotion
    products = Product.objects.select_related('category').all()
    categories = Category.objects.all()
    active_promotions = Promotion.objects.filter(active=True)
    quick_products = (
        Product.objects.filter(name__icontains='pain')
        | Product.objects.filter(name__icontains='glace')
        | Product.objects.filter(category__name__icontains='pain')
        | Product.objects.filter(category__name__icontains='glace')
    ).distinct()[:8]
    settings = StoreSettings.load()
    return render(request, 'store/pos.html', {
        'products': products,
        'categories': categories,
        'active_promotions': active_promotions,
        'quick_products': quick_products,
        'default_currency': settings.currency,
    })


# ===== CLIENTS =====
def customer_list(request):
    customers = Customer.objects.annotate(total_spent=Sum('sales__total'), sale_count=Count('sales')).order_by('-sale_count')
    return render(request, 'store/customer_list.html', {'customers': customers})


def customer_history(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    sales = Sale.objects.filter(customer=customer).order_by('-sale_date')
    return render(request, 'store/customer_history.html', {'customer': customer, 'sales': sales})


# ===== TOP PRODUITS =====
def top_products(request):
    from django.db.models import Sum, F
    top = (SaleItem.objects.values('product__id', 'product__name', 'product__price')
           .annotate(total_qty=Sum('quantity'))
           .order_by('-total_qty')[:20])
    return render(request, 'store/top_products.html', {'top_products': top})


# ===== INVENTAIRE =====
def inventory_list(request):
    from .forms import InventoryAdjustmentForm
    from .models import InventoryAdjustment
    form = InventoryAdjustmentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        adj = form.save(commit=False)
        adj.system_stock = adj.product.stock
        adj.difference = adj.counted_stock - adj.system_stock
        adj.save()
        adj.product.stock = adj.counted_stock
        adj.product.save(update_fields=['stock'])
        messages.success(request, f'Inventaire mis à jour : {adj.product.name} (écart de {adj.difference})')
        return redirect('store:inventory_list')
    adjustments = InventoryAdjustment.objects.select_related('product')[:50]
    return render(request, 'store/inventory_list.html', {'form': form, 'adjustments': adjustments})


# ===== EXPORT EXCEL =====
def export_sales_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventes'
    headers = ['ID', 'Date', 'Client', 'Total', 'Payé', 'Statut', 'Méthode']
    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='2563EB')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = fill; cell.alignment = Alignment(horizontal='center')
    for i, sale in enumerate(Sale.objects.all(), 2):
        ws.cell(row=i, column=1, value=sale.id)
        ws.cell(row=i, column=2, value=str(sale.sale_date))
        ws.cell(row=i, column=3, value=sale.customer_name or '-')
        ws.cell(row=i, column=4, value=float(sale.total))
        ws.cell(row=i, column=5, value=float(sale.amount_paid))
        ws.cell(row=i, column=6, value=sale.payment_status)
        ws.cell(row=i, column=7, value=sale.payment_method or '-')
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventes.xlsx'
    wb.save(response)
    return response


def export_products_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Produits'
    headers = ['Code', 'Nom', 'Catégorie', 'Prix vente', "Prix d'achat", 'Stock', 'Stock min']
    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='16A34A')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = fill; cell.alignment = Alignment(horizontal='center')
    for i, p in enumerate(Product.objects.all(), 2):
        ws.cell(row=i, column=1, value=p.code)
        ws.cell(row=i, column=2, value=p.name)
        ws.cell(row=i, column=3, value=p.category.name if p.category else '')
        ws.cell(row=i, column=4, value=float(p.price))
        ws.cell(row=i, column=5, value=float(p.cost_price))
        ws.cell(row=i, column=6, value=float(p.stock))
        ws.cell(row=i, column=7, value=float(p.min_stock) if p.min_stock else 0)
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=produits.xlsx'
    wb.save(response)
    return response


# ===== ACTIONS GROUPÉES PRODUITS =====
def product_batch_action(request):
    if request.method != 'POST':
        return redirect('store:product_manage')
    action = request.POST.get('action', '')
    ids = request.POST.getlist('ids[]')
    if not ids:
        messages.warning(request, 'Aucun produit sélectionné.')
        return redirect('store:product_manage')
    products = Product.objects.filter(pk__in=ids)
    count = products.count()
    if action == 'delete':
        deleted, _ = products.delete()
        messages.success(request, f'{deleted} produit(s) supprimé(s).')
    elif action == 'category':
        cat_id = request.POST.get('category_id', '')
        if cat_id:
            updated = products.update(category_id=cat_id)
            messages.success(request, f'{updated} produit(s) mis à jour.')
        else:
            messages.error(request, 'Catégorie non spécifiée.')
    elif action == 'price':
        try:
            new_price = Decimal(request.POST.get('new_price', '0'))
            price_type = request.POST.get('price_type', 'price')
            kwargs = {price_type: new_price}
            updated = products.update(**kwargs)
            messages.success(request, f'{updated} produit(s) mis à jour.')
        except (InvalidOperation, ValueError):
            messages.error(request, 'Prix invalide.')
    else:
        messages.error(request, 'Action inconnue.')
    return redirect('store:product_manage')


# ===== IMPORT CSV/EXCEL PRODUITS =====
def product_import_bulk(request):
    if request.method != 'POST':
        return redirect('store:product_manage')
    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('store:product_manage')
    ext = os.path.splitext(file.name)[1].lower()
    if ext not in ('.csv', '.xlsx', '.xls'):
        messages.error(request, 'Format supporté : CSV ou Excel (.xlsx, .xls).')
        return redirect('store:product_manage')
    imported = errors = 0
    try:
        if ext == '.csv':
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(StringIO(content))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        for row in rows:
            try:
                name = (row.get('Nom') or row.get('name') or '').strip()
                if not name:
                    errors += 1
                    continue
                category_name = (row.get('Catégorie') or row.get('category') or '').strip()
                category = None
                if category_name:
                    category, _ = Category.objects.get_or_create(name=category_name)
                barcode_val = str(row.get('Code-barres') or row.get('barcode') or '').strip()
                code_val = str(row.get('Code') or row.get('code') or '').strip()
                if not code_val:
                    code_val = None
                Product.objects.create(
                    code=code_val or None,
                    barcode=barcode_val or None,
                    name=name,
                    category=category,
                    unit=(row.get('Unité') or row.get('unit') or 'piece').strip(),
                    price=Decimal(str(row.get('Prix vente') or row.get('price') or row.get('Prix') or 0)),
                    cost_price=Decimal(str(row.get("Prix d'achat") or row.get('cost_price') or 0)),
                    stock=int(float(str(row.get('Stock') or row.get('stock') or 0))),
                    min_stock=int(float(str(row.get('Stock min') or row.get('min_stock') or 0))),
                )
                imported += 1
            except Exception:
                errors += 1
    except Exception as e:
        messages.error(request, f'Erreur fichier : {e}')
        return redirect('store:product_manage')
    msg = f'{imported} produit(s) importé(s) avec succès.'
    if errors:
        msg += f' {errors} ligne(s) ignorée(s) (erreurs).'
    messages.success(request, msg)
    return redirect('store:product_manage')


# ===== PRODUITS EN RUPTURE =====
def out_of_stock(request):
    products = Product.objects.filter(stock=0).order_by('name')
    return render(request, 'store/out_of_stock.html', {'products': products})


# ===== DUPLICATION PRODUIT =====
def product_duplicate(request, pk):
    original = get_object_or_404(Product, pk=pk)
    original.pk = None
    original.code = ''
    original.name = original.name + ' (copie)'
    original.save()
    messages.success(request, f'Produit dupliqué. Modifie-le maintenant.')
    return redirect('store:product_update', pk=original.pk)


# ===== RAPPORT PAIEMENTS =====
def payment_method_report(request):
    from django.db.models import Count, Sum
    data = Sale.objects.values('payment_method').annotate(
        count=Count('id'), total=Sum('total')
    ).order_by('-total')
    return render(request, 'store/payment_report.html', {'data': data})


# ===== SAUVEGARDE AUTO =====
def database_auto_backup(request):
    try:
        import shutil, datetime
        db_path = None
        for alias in settings.DATABASES:
            opt = settings.DATABASES[alias]
            if 'NAME' in opt and opt['NAME'].endswith('.sqlite3'):
                db_path = opt['NAME']
                break
        if db_path:
            backup_dir = os.path.join(os.path.dirname(db_path), 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            shutil.copy2(db_path, os.path.join(backup_dir, f'backup_{ts}.sqlite3'))
            messages.success(request, f'Sauvegarde effectuée : backup_{ts}.sqlite3')
        else:
            messages.error(request, 'Base de données non trouvée.')
    except Exception as e:
        messages.error(request, f'Erreur sauvegarde : {e}')
    return redirect('store:database_tools')


# ===== RECETTES / PRODUITS COMPOSITES (BOM) =====
def recipe_list(request):
    recipes = Recipe.objects.select_related('product').prefetch_related('ingredients__product').all()
    return render(request, 'store/recipe_list.html', {'recipes': recipes})


def recipe_create(request):
    products = Product.objects.filter(is_composite=False).order_by('name')
    if request.method == 'POST':
        product_id = request.POST.get('product')
        ingredient_ids = request.POST.getlist('ingredient')
        quantities = request.POST.getlist('quantity')
        if not product_id:
            messages.error(request, 'Sélectionne un produit composite.')
        elif not ingredient_ids or not any(q.strip() for q in quantities):
            messages.error(request, 'Ajoute au moins un ingrédient.')
        else:
            product = get_object_or_404(Product, pk=product_id)
            try:
                with transaction.atomic():
                    recipe = Recipe.objects.create(product=product)
                    product.is_composite = True
                    product.save(update_fields=['is_composite'])
                    for ing_id, qty in zip(ingredient_ids, quantities):
                        qty = qty.strip()
                        if not qty or not ing_id:
                            continue
                        ing_product = Product.objects.get(pk=ing_id)
                        RecipeIngredient.objects.create(
                            recipe=recipe, product=ing_product,
                            quantity=Decimal(qty)
                        )
                    messages.success(request, f'Recette créée pour {product.name}')
                    return redirect('store:recipe_list')
            except Exception as e:
                messages.error(request, f'Erreur : {e}')
    return render(request, 'store/recipe_form.html', {
        'products': products,
        'ingredients': Product.objects.filter(is_composite=False).order_by('name'),
    })


def recipe_edit(request, pk):
    recipe = get_object_or_404(Recipe.objects.select_related('product'), pk=pk)
    if request.method == 'POST':
        ingredient_ids = request.POST.getlist('ingredient')
        quantities = request.POST.getlist('quantity')
        with transaction.atomic():
            recipe.ingredients.all().delete()
            for ing_id, qty in zip(ingredient_ids, quantities):
                qty = qty.strip()
                if not qty or not ing_id:
                    continue
                ing_product = Product.objects.get(pk=ing_id)
                RecipeIngredient.objects.create(
                    recipe=recipe, product=ing_product,
                    quantity=Decimal(qty)
                )
            messages.success(request, 'Recette mise à jour.')
            return redirect('store:recipe_list')
    return render(request, 'store/recipe_form.html', {
        'recipe': recipe,
        'products': Product.objects.filter(is_composite=False).order_by('name'),
        'ingredients': Product.objects.filter(is_composite=False).order_by('name'),
    })


def recipe_delete(request, pk):
    recipe = get_object_or_404(Recipe, pk=pk)
    product = recipe.product
    if request.method == 'POST':
        product.is_composite = False
        product.save(update_fields=['is_composite'])
        recipe.delete()
        messages.success(request, 'Recette supprimée.')
        return redirect('store:recipe_list')
    return render(request, 'store/recipe_confirm_delete.html', {'recipe': recipe})


# ===== RAPPORT VALORISATION STOCK =====
def stock_valuation_report(request):
    products = Product.objects.select_related('category').all().order_by('name')
    total_value = sum(p.stock * p.cost_price for p in products if p.cost_price > 0)
    total_sale_value = sum(p.stock * p.price for p in products)
    return render(request, 'store/stock_valuation.html', {
        'products': products,
        'total_value': total_value,
        'total_sale_value': total_sale_value,
    })


# ===== EMPLOYÉS =====
def employee_list(request):
    employees = Employee.objects.all().order_by('last_name', 'first_name')
    return render(request, 'store/employee_list.html', {'employees': employees})


def employee_create(request):
    if request.method == 'POST':
        form = EmployeeCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Employé ajouté.')
            return redirect('store:employee_list')
    else:
        form = EmployeeCreateForm()
    return render(request, 'store/employee_form.html', {'form': form, 'title': 'Nouvel employé'})


def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeCreateForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, 'Employé modifié.')
            return redirect('store:employee_list')
    else:
        form = EmployeeCreateForm(instance=employee)
    return render(request, 'store/employee_form.html', {'form': form, 'title': 'Modifier employé', 'employee': employee})


def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'Employé supprimé.')
        return redirect('store:employee_list')
    return render(request, 'store/employee_confirm_delete.html', {'employee': employee})


def attendance_list(request):
    attendances = Attendance.objects.select_related('employee').all().order_by('-clock_in')[:50]
    employees = Employee.objects.filter(is_active=True)
    today = date.today()
    return render(request, 'store/attendance_list.html', {
        'attendances': attendances,
        'employees': employees,
        'today': today,
    })


def attendance_clock_in(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if Attendance.objects.filter(employee=employee, clock_out__isnull=True).exists():
        messages.warning(request, f'{employee} est déjà en service.')
    else:
        Attendance.objects.create(employee=employee)
        messages.success(request, f'{employee} a pointé son arrivée.')
    return redirect('store:attendance_list')


def attendance_clock_out(request, pk):
    att = get_object_or_404(Attendance, pk=pk, clock_out__isnull=True)
    att.clock_out = timezone.localtime()
    att.save(update_fields=['clock_out'])
    messages.success(request, f'{att.employee} a pointé son départ ({att.duration}h).')
    return redirect('store:attendance_list')


# ===== PAIE / RH =====
def payroll_report(request):
    employees = Employee.objects.filter(is_active=True)
    month = int(request.GET.get('month', date.today().month))
    year = int(request.GET.get('year', date.today().year))

    report_data = []
    total_salary = Decimal('0')
    for emp in employees:
        atts = emp.attendance.filter(
            clock_in__year=year, clock_in__month=month,
            clock_out__isnull=False
        )
        total_hours = sum((a.duration or 0) for a in atts)
        days_worked = atts.count()
        # prorata : salaire mensuel / 30 * jours travaillés
        prorata = (emp.salary / Decimal('30')) * Decimal(str(days_worked)) if days_worked else Decimal('0')
        report_data.append({
            'employee': emp,
            'days_worked': days_worked,
            'total_hours': round(total_hours, 1),
            'prorata': prorata,
        })
        total_salary += prorata

    months = [(i, date(year, i, 1).strftime('%B')) for i in range(1, 13)]
    years = list(range(date.today().year - 2, date.today().year + 1))

    return render(request, 'store/payroll_report.html', {
        'report_data': report_data,
        'month': month,
        'year': year,
        'months': months,
        'years': years,
        'total_salary': total_salary,
    })


# ===== RAPPORT MARGE DÉTAILLÉ =====
def margin_report(request):
    months_back = int(request.GET.get('months', 6))
    now = date.today()
    start = now.replace(month=1, day=1) if months_back == 0 else now - timedelta(days=30 * months_back)

    sales = Sale.objects.filter(sale_date__gte=start).prefetch_related('items__product')
    category_data = {}
    monthly_data = {}
    total_revenue = Decimal('0')
    total_cost = Decimal('0')

    PACK_MODES = {'paquet': 1, 'carton': 12, 'cartouche': 1}

    for sale in sales:
        mkey = sale.sale_date.strftime('%Y-%m')
        if mkey not in monthly_data:
            monthly_data[mkey] = {'revenue': Decimal('0'), 'cost': Decimal('0')}

        for item in sale.items.all():
            if item.sale_mode in PACK_MODES and item.product.pack_size > 1:
                mult = PACK_MODES[item.sale_mode]
                effective_qty = item.quantity * item.product.pack_size * mult
            else:
                effective_qty = item.quantity
            revenue = item.subtotal
            cost = effective_qty * (item.product.cost_price or 0)

            total_revenue += revenue
            total_cost += cost
            monthly_data[mkey]['revenue'] += revenue
            monthly_data[mkey]['cost'] += cost

            cat = item.product.category
            cat_name = cat.name if cat else 'Sans catégorie'
            if cat_name not in category_data:
                category_data[cat_name] = {'revenue': Decimal('0'), 'cost': Decimal('0'), 'count': 0}
            category_data[cat_name]['revenue'] += revenue
            category_data[cat_name]['cost'] += cost
            category_data[cat_name]['count'] += 1

    for cat in category_data.values():
        cat['profit'] = cat['revenue'] - cat['cost']
        cat['margin'] = float(cat['profit'] / cat['revenue'] * 100) if cat['revenue'] else 0

    months_list = sorted(monthly_data.items())
    for m in months_list:
        m[1]['profit'] = m[1]['revenue'] - m[1]['cost']
        m[1]['margin'] = float(m[1]['profit'] / m[1]['revenue'] * 100) if m[1]['revenue'] else 0

    categories = sorted(category_data.items(), key=lambda x: -x[1]['profit'])
    total_profit = total_revenue - total_cost
    total_margin = float(total_profit / total_revenue * 100) if total_revenue else 0

    return render(request, 'store/margin_report.html', {
        'categories': categories,
        'months': months_list,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'total_margin': total_margin,
        'months_back': months_back,
        'start': start,
        'end': now,
    })


# ===== DASHBOARD RH =====
def rh_dashboard(request):
    today = date.today()
    employees = Employee.objects.all()
    total = employees.count()
    active = employees.filter(is_active=True).count()

    today_attendance = Attendance.objects.filter(
        clock_in__date=today
    ).select_related('employee').order_by('-clock_in')

    clocked_in = today_attendance.filter(clock_out__isnull=True).count()
    clocked_out = today_attendance.exclude(clock_out__isnull=True).count()

    month_start = today.replace(day=1)
    month_attendance = Attendance.objects.filter(
        clock_in__date__gte=month_start, clock_in__date__lte=today,
        clock_out__isnull=False
    )
    total_work_hours = sum((a.duration or 0) for a in month_attendance)
    working_days = month_attendance.dates('clock_in', 'day').count()
    attendance_rate = round((working_days / (today.day or 1)) * 100, 1) if today.day else 0

    total_payroll = sum(
        (emp.salary / Decimal('30')) * Decimal(str(
            month_attendance.filter(employee=emp, clock_out__isnull=False).dates('clock_in', 'day').count()
        ))
        for emp in employees.filter(is_active=True)
    )

    return render(request, 'store/rh_dashboard.html', {
        'total_employees': total,
        'active_employees': active,
        'clocked_in': clocked_in,
        'clocked_out': clocked_out,
        'today_attendance': today_attendance,
        'total_work_hours': round(total_work_hours, 1),
        'working_days': working_days,
        'today': today,
        'attendance_rate': attendance_rate,
        'total_payroll': total_payroll,
    })


# ===== PDF MANUEL D'UTILISATION =====
def user_manual_pdf(request):
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, PageBreak,
        Table, TableStyle, ListFlowable, ListItem,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus.flowables import HRFlowable

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=25*mm, bottomMargin=20*mm,
        leftMargin=20*mm, rightMargin=20*mm,
    )

    styles = getSampleStyleSheet()
    style_h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, spaceAfter=12, textColor=rl_colors.HexColor('#7f1d1d'))
    style_h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=15, spaceAfter=8, spaceBefore=16, textColor=rl_colors.HexColor('#7f1d1d'))
    style_h3 = ParagraphStyle('H3', parent=styles['Heading3'], fontName='Helvetica-Bold', fontSize=12, spaceAfter=6, spaceBefore=10, textColor=rl_colors.HexColor('#334155'))
    style_body = ParagraphStyle('Body', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, spaceAfter=6)
    style_bullet = ParagraphStyle('Bullet', parent=style_body, leftIndent=18, bulletIndent=6, spaceAfter=3)
    style_code = ParagraphStyle('Code', parent=styles['Code'], fontName='Courier', fontSize=8, spaceAfter=4, leftIndent=10, backColor=rl_colors.HexColor('#f1f5f9'))

    story = []

    # ===== COVER PAGE =====
    story.append(Spacer(1, 100*mm))
    story.append(Paragraph('MANUEL D\'UTILISATION', ParagraphStyle('CoverTitle', fontName='Helvetica-Bold', fontSize=32, textColor=rl_colors.HexColor('#7f1d1d'), alignment=TA_CENTER)))
    story.append(Spacer(1, 12*mm))
    story.append(Paragraph('Logiciel de Gestion de Supermarché', ParagraphStyle('CoverSub', fontName='Helvetica', fontSize=16, textColor=rl_colors.HexColor('#475569'), alignment=TA_CENTER)))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(f'Version {getattr(settings, "APP_VERSION", "1.0")}', ParagraphStyle('CoverVer', fontName='Helvetica', fontSize=11, textColor=rl_colors.HexColor('#94a3b8'), alignment=TA_CENTER)))
    story.append(Spacer(1, 30*mm))
    story.append(HRFlowable(width='60%', thickness=2, color=rl_colors.HexColor('#7f1d1d')))
    story.append(Spacer(1, 8*mm))
    from datetime import datetime
    story.append(Paragraph(f'Généré le {datetime.now().strftime("%d/%m/%Y à %H:%M")}', ParagraphStyle('CoverDate', fontName='Helvetica', fontSize=10, textColor=rl_colors.HexColor('#94a3b8'), alignment=TA_CENTER)))
    story.append(PageBreak())

    # ===== TABLE OF CONTENTS =====
    story.append(Paragraph('Table des matières', style_h1))
    story.append(Spacer(1, 6*mm))
    toc_items = [
        '1. Présentation générale',
        '2. Tableau de bord',
        '3. Gestion des produits',
        '4. Ventes et caisse',
        '5. Achats et approvisionnements',
        '6. Dépenses et dettes',
        '7. Clients et crédits téléphoniques',
        '8. Caisse et sessions',
        '9. Retours fournisseurs',
        '10. Promotions',
        '11. Recettes / Produits composites',
        '12. Employés et pointage',
        '13. Paie et RH',
        '14. Rapports et marges',
        '15. Paramètres et configuration',
        '16. Sauvegarde et base de données',
        '17. Raccourcis clavier',
    ]
    for item in toc_items:
        story.append(Paragraph(item, ParagraphStyle('TOC', fontName='Helvetica', fontSize=11, leading=18, textColor=rl_colors.HexColor('#334155'))))
    story.append(PageBreak())

    # ===== SECTION 1 =====
    story.append(Paragraph('1. Présentation générale', style_h1))
    story.append(Paragraph(
        'Ce logiciel de gestion de supermarché permet de gérer l\'intégralité des opérations d\'un point de vente : '
        'de la réception des marchandises à la vente au comptoir, en passant par le suivi des stocks, des employés, '
        'des finances et de la relation client. L\'interface est conçue pour être utilisée aussi bien sur ordinateur '
        'que sur tablette et téléphone mobile.',
        style_body
    ))
    story.append(Paragraph('Fonctionnalités principales', style_h2))
    for f in [
        'Caisse tactile avec scan de code-barres',
        'Gestion complète des stocks et inventaire',
        'Module RH avec pointage et paie',
        'Recettes / nomenclature (BOM) pour produits composites',
        'Rapports détaillés et valorisation du stock',
        'Sauvegarde automatique de la base de données',
        'Interface responsive (mobile, tablette, desktop)',
        'Thèmes d\'interface personnalisables',
    ]:
        story.append(Paragraph(f'• {f}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 2 =====
    story.append(Paragraph('2. Tableau de bord', style_h1))
    story.append(Paragraph(
        'La page d\'accueil (<b>/</b>) affiche un tableau de bord avec des tuiles d\'accès rapide '
        'aux différentes fonctionnalités du logiciel. Les tuiles sont organisées par catégorie : '
        'Caisse, Produits, Stock, Approvisionnements, Clients, Historique, Dépenses, Dettes, Bénéfices, etc.',
        style_body
    ))
    story.append(Paragraph('Barre de navigation latérale', style_h2))
    story.append(Paragraph(
        'La barre latérale gauche permet d\'accéder à toutes les sections. '
        'Elle peut être réduite (mini-sidebar) à l\'aide du bouton hamburger ☰. '
        'Sur mobile, elle se cache automatiquement et s\'affiche par-dessus le contenu.',
        style_body
    ))
    story.append(Paragraph('Barre du haut', style_h2))
    story.append(Paragraph(
        'La barre d\'en-tête contient le nom du magasin, la barre de recherche globale, '
        'le statut réseau (en ligne/hors ligne), le compteur de notifications, '
        'le compteur du panier et le menu utilisateur.',
        style_body
    ))
    story.append(Paragraph('Navigation mobile', style_h2))
    story.append(Paragraph(
        'Sur téléphone, une barre de navigation inférieure fixe donne accès aux 5 sections '
        'principales : Accueil, Vente, Panier, Produits, Plus.',
        style_body
    ))
    story.append(PageBreak())

    # ===== SECTION 3 =====
    story.append(Paragraph('3. Gestion des produits', style_h1))
    story.append(Paragraph(
        'La section Produits permet de créer, modifier, supprimer et gérer le stock de tous les articles du magasin.',
        style_body
    ))
    story.append(Paragraph('Créer un produit', style_h2))
    for step in [
        'Aller dans Produits → Gérer → Nouveau produit',
        'Renseigner le nom, le code (automatique si vide), le code-barres (scannable)',
        'Sélectionner la catégorie ou créer une nouvelle catégorie',
        'Définir l\'unité de vente (pièce, kg, litre, etc.)',
        'Saisir le prix de vente et le prix d\'achat (coût)',
        'Définir le stock initial et le stock minimum d\'alerte',
        'Optionnel : taille du paquet, prix du paquet, taille du carton',
        'Optionnel : date de péremption',
        'Enregistrer',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Gestion du stock', style_h2))
    story.append(Paragraph(
        'Depuis la liste des produits, vous pouvez : approvisionner (augmenter le stock), '
        'effectuer un stock rapide, voir l\'historique des prix, dupliquer un produit, '
        'et imprimer des étiquettes de prix avec code-barres.',
        style_body
    ))
    story.append(Paragraph('Import/Export', style_h2))
    story.append(Paragraph(
        'Les produits peuvent être importés en masse depuis un fichier Excel et exportés '
        'vers Excel ou CSV. Des actions par lot (modification de prix, changement de catégorie) '
        'sont disponibles dans la vue de gestion.',
        style_body
    ))
    story.append(PageBreak())

    # ===== SECTION 4 =====
    story.append(Paragraph('4. Ventes et caisse', style_h1))
    story.append(Paragraph(
        'Le module de vente gère l\'encaissement des produits avec plusieurs modes de vente.',
        style_body
    ))
    story.append(Paragraph('Vente simple', style_h2))
    for step in [
        'Aller dans Caisse ou Vente',
        'Scanner le code-barres d\'un produit ou le sélectionner dans la liste',
        'Saisir la quantité et le mode (pièce, paquet, carton)',
        'Le total est calculé automatiquement',
        'Choisir le mode de paiement (espèces, Orange Money, etc.)',
        'Valider la vente',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Caisse tactile (POS)', style_h2))
    story.append(Paragraph(
        'La vue Caisse tactile (<b>/caisse/</b>) est optimisée pour les écrans tactiles avec '
        'des grandes tuiles produits. Un scan de code-barres intégré (via caméra) permet '
        'd\'ajouter rapidement les produits. Le panier s\'affiche à droite et les totaux '
        'sont mis à jour en temps réel.',
        style_body
    ))
    story.append(Paragraph('Panier', style_h2))
    story.append(Paragraph(
        'Le panier permet d\'accumuler des produits avant de finaliser une vente. '
        'Accessible via l\'icône panier dans l\'en-tête ou le bouton flottant sur mobile. '
        'Vous pouvez ajouter/supprimer des articles depuis la page produit.',
        style_body
    ))
    story.append(Paragraph('Historique et détail des ventes', style_h2))
    story.append(Paragraph(
        'Toutes les ventes sont enregistrées avec date, heure, total, mode de paiement '
        'et statut de paiement. Vous pouvez visualiser, modifier ou supprimer une vente, '
        'et imprimer une facture ou un ticket.',
        style_body
    ))
    story.append(PageBreak())

    # ===== SECTION 5 =====
    story.append(Paragraph('5. Achats et approvisionnements', style_h1))
    story.append(Paragraph(
        'Le module d\'approvisionnement permet de gérer les entrées de stock.',
        style_body
    ))
    story.append(Paragraph('Approvisionnement direct', style_h2))
    for step in [
        'Depuis la fiche produit, cliquer sur Approvisionner',
        'Saisir la quantité ajoutée et le prix d\'achat',
        'Le stock est automatiquement augmenté',
        'L\'historique des prix d\'achat est conservé',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Réapprovisionnement automatique', style_h2))
    story.append(Paragraph(
        'La vue Réassort automatique liste les produits dont le stock est inférieur au seuil minimum. '
        'Vous pouvez générer une commande fournisseur en un clic.',
        style_body
    ))
    story.append(Paragraph('Commandes fournisseurs', style_h2))
    story.append(Paragraph(
        'Les bons de commande (Purchase Orders) permettent de passer commande auprès des fournisseurs. '
        'À la réception, vous validez la commande et le stock est mis à jour automatiquement.',
        style_body
    ))
    story.append(PageBreak())

    # ===== SECTION 6 =====
    story.append(Paragraph('6. Dépenses et dettes', style_h1))
    story.append(Paragraph(
        'Suivez les dépenses courantes du magasin (loyer, électricité, salaires, etc.) '
        'ainsi que les dettes actives et passives.',
        style_body
    ))
    story.append(Paragraph('Dépenses', style_h2))
    for step in [
        'Aller dans Dépenses → Nouvelle dépense',
        'Saisir le montant, la catégorie et la date',
        'Optionnel : ajouter une note',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Dettes', style_h2))
    for step in [
        'Aller dans Dettes → Nouvelle dette',
        'Saisir le contact, le montant, la date d\'échéance',
        'Marquer comme réglée une fois remboursée',
        'Les dettes en retard sont surlignées en rouge',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 7 =====
    story.append(Paragraph('7. Clients et crédits téléphoniques', style_h1))
    story.append(Paragraph(
        'Gérez vos clients et leur historique d\'achats, ainsi que la vente de crédits téléphoniques.',
        style_body
    ))
    story.append(Paragraph('Clients', style_h2))
    for step in [
        'Aller dans Clients pour voir la liste',
        'Cliquer sur un client pour voir son historique d\'achats',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Crédits téléphoniques', style_h2))
    for step in [
        'Configurer les opérateurs et montants dans Crédits téléphoniques',
        'Vendre un crédit depuis la caisse ou depuis le module dédié',
        'Suivre les achats de crédits dans l\'historique',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 8 =====
    story.append(Paragraph('8. Caisse et sessions', style_h1))
    story.append(Paragraph(
        'Le module Caisse permet d\'ouvrir et fermer des sessions de caisse pour suivre '
        'les encaissements et les mouvements d\'argent.',
        style_body
    ))
    for step in [
        'Ouvrir une session avec le montant de départ',
        'Les ventes sont automatiquement enregistrées dans la session en cours',
        'Fermer la session : le solde final est calculé automatiquement',
        'Un rapport de clôture récapitule les ventes et mouvements',
        'Ajouter des mouvements (retrait/dépôt) via Caisse → Mouvement',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 9 =====
    story.append(Paragraph('9. Retours fournisseurs', style_h1))
    story.append(Paragraph(
        'Gérez les retours de produits aux fournisseurs (produits défectueux, périmés, etc.) '
        'avec suivi des remboursements.',
        style_body
    ))
    for step in [
        'Aller dans Retours → Nouveau retour',
        'Sélectionner le fournisseur et ajouter les produits retournés',
        'Le montant du remboursement est calculé automatiquement',
        'Approuver ou rejeter le retour',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 10 =====
    story.append(Paragraph('10. Promotions', style_h1))
    story.append(Paragraph(
        'Créez des promotions sur les produits pour une période donnée.',
        style_body
    ))
    for step in [
        'Aller dans Promotions → Nouvelle promotion',
        'Sélectionner un produit, définir le prix promotionnel',
        'Définir les dates de début et fin',
        'Activer/désactiver la promotion',
        'Les promotions actives sont affichées dans la caisse',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 11 =====
    story.append(Paragraph('11. Recettes / Produits composites (BOM)', style_h1))
    story.append(Paragraph(
        'Le module Recettes permet de définir des produits composites fabriqués à partir '
        'd\'ingrédients. Lors de la vente d\'un produit composite, le stock de ses ingrédients '
        'est automatiquement décrémenté.',
        style_body
    ))
    story.append(Paragraph('Créer une recette', style_h2))
    for step in [
        'Aller dans Produits → Recettes → Nouvelle recette',
        'Sélectionner le produit composite (ex: "Assiette poulet")',
        'Ajouter les ingrédients avec leurs quantités (ex: 1 poulet, 2 kg riz, 0.5 L huile)',
        'Le coût total et la marge sont calculés automatiquement',
        'Enregistrer',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Vente d\'un produit composite', style_h2))
    story.append(Paragraph(
        'Lors de la vente (Caisse ou Vente simple), les ingrédients du produit composite '
        'sont automatiquement déduits du stock. Le détail de la consommation apparaît '
        'dans la page de détail de la vente.',
        style_body
    ))
    story.append(Paragraph('Traçabilité', style_h2))
    story.append(Paragraph(
        'Dans le détail d\'une vente, les produits composites affichent la liste des '
        'ingrédients consommés avec les quantités correspondantes.',
        style_body
    ))
    story.append(PageBreak())

    # ===== SECTION 12 =====
    story.append(Paragraph('12. Employés et pointage', style_h1))
    story.append(Paragraph(
        'Le module RH permet de gérer les employés et leurs pointages (arrivée/départ).',
        style_body
    ))
    story.append(Paragraph('Gestion des employés', style_h2))
    for step in [
        'Aller dans Ressources → Employés → Nouvel employé',
        'Saisir le nom, prénom, téléphone, email',
        'Choisir le poste (caissier, gérant, magasinier, etc.)',
        'Définir le salaire mensuel et la date d\'embauche',
        'Activer ou désactiver l\'employé',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Pointage (Arrivée/Départ)', style_h2))
    for step in [
        'Aller dans Ressources → Pointage',
        'Voir les employés actifs avec leur statut (Absent / En service)',
        'Cliquer sur "Pointer arrivée" ou "Pointer départ"',
        'Les heures travaillées sont calculées automatiquement',
        'L\'historique des 50 derniers pointages est affiché',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 13 =====
    story.append(Paragraph('13. Paie et RH', style_h1))
    story.append(Paragraph(
        'Le tableau de bord RH et le rapport de paie synthétisent les données des employés '
        'et des pointages.',
        style_body
    ))
    story.append(Paragraph('Dashboard RH', style_h2))
    for step in [
        'Accessible via /rh/ ou depuis le menu RH',
        'Affiche : nombre d\'employés, en service aujourd\'hui, heures du mois, masse salariale',
        'Liste des employés avec leur statut du jour',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Rapport de paie', style_h2))
    for step in [
        'Accessible via /paie/ ou depuis le menu Paie',
        'Sélectionner le mois et l\'année',
        'Le rapport calcule pour chaque employé : jours travaillés, heures totales, salaire prorata',
        'La masse salariale totale est affichée en haut',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 14 =====
    story.append(Paragraph('14. Rapports et marges', style_h1))
    story.append(Paragraph(
        'Plusieurs rapports sont disponibles pour analyser l\'activité du magasin.',
        style_body
    ))
    story.append(Paragraph('Rapport journalier', style_h2))
    story.append(Paragraph(
        'Affiche les ventes du jour, les dépenses, les dettes échues, les produits en rupture '
        'et les performances de la journée.',
        style_body
    ))
    story.append(Paragraph('Rapport de marges', style_h2))
    story.append(Paragraph(
        'Analyse détaillée des marges par catégorie de produits avec évolution mensuelle. '
        'Sélectionnez la période (3, 6, 12 mois ou année en cours).',
        style_body
    ))
    story.append(Paragraph('Valorisation du stock', style_h2))
    story.append(Paragraph(
        'Calcule la valeur totale du stock au coût d\'achat et au prix de vente. '
        'Utile pour les inventaires comptables.',
        style_body
    ))
    story.append(Paragraph('Rapport par moyen de paiement', style_h2))
    story.append(Paragraph(
        'Répartition des ventes par moyen de paiement (espèces, Orange Money, Mobile Money, etc.) '
        'avec le nombre de transactions et les totaux.',
        style_body
    ))
    story.append(Paragraph('Rapport de bénéfices', style_h2))
    story.append(Paragraph(
        'Analyse des bénéfices par produit sur une période donnée avec calcul de la marge brute.',
        style_body
    ))
    story.append(PageBreak())

    # ===== SECTION 15 =====
    story.append(Paragraph('15. Paramètres et configuration', style_h1))
    story.append(Paragraph(
        'La page Paramètres permet de configurer l\'ensemble du logiciel.',
        style_body
    ))
    story.append(Paragraph('Paramètres disponibles', style_h2))
    params = [
        'Nom du magasin et message d\'accueil',
        'Logo et image de fond',
        'Thème d\'interface (Bordeaux, Bleu, Vert, Sombre, Clair, Cyber futuriste)',
        'Disposition des factures (Classique, Compacte, Moderne)',
        'Devise (FCFA, EUR, USD), taux de change',
        'Adresse, téléphone, signature',
        'Alertes vocales',
        'Limite de dépenses mensuelles',
        'Configuration SMS et WhatsApp',
        'Compte bancaire par défaut',
        'Format des étiquettes de prix',
    ]
    for p in params:
        story.append(Paragraph(f'• {p}', style_bullet))
    story.append(PageBreak())

    # ===== SECTION 16 =====
    story.append(Paragraph('16. Sauvegarde et base de données', style_h1))
    story.append(Paragraph(
        'Le logiciel inclut plusieurs outils pour la sauvegarde et la maintenance '
        'de la base de données.',
        style_body
    ))
    story.append(Paragraph('Sauvegardes', style_h2))
    for step in [
        'Sauvegarde manuelle : Outils → Base de données → Sauvegarder',
        'Sauvegarde automatique programmée (rotative, garde les 20 dernières)',
        'Export JSON complet de la base',
        'Import JSON pour restaurer une sauvegarde',
        'Nettoyage des anciennes données',
    ]:
        story.append(Paragraph(f'• {step}', style_bullet))

    story.append(Paragraph('Les sauvegardes sont stockées dans le dossier <b>backups/</b> '
                           'au même niveau que la base de données.', style_body))
    story.append(PageBreak())

    # ===== SECTION 17 =====
    story.append(Paragraph('17. Raccourcis clavier', style_h1))
    story.append(Paragraph(
        'Des raccourcis clavier sont disponibles pour accélérer la navigation '
        '(accessibles via ? depuis n\'importe quelle page).',
        style_body
    ))
    shortcuts = [
        ['<b>Échap</b>', 'Fermer la modale / revenir en arrière'],
        ['<b>?</b>', 'Afficher l\'aide des raccourcis'],
        ['<b>Ctrl+F</b>', 'Rechercher un produit'],
        ['<b>Ctrl+N</b>', 'Nouvelle vente'],
        ['<b>/</b>', 'Focus barre de recherche'],
    ]
    for k, v in shortcuts:
        story.append(Paragraph(f'{k} — {v}', style_bullet))

    story.append(Spacer(1, 20*mm))
    story.append(HRFlowable(width='100%', thickness=1, color=rl_colors.HexColor('#e2e8f0')))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        'Fin du manuel — Merci d\'utiliser ce logiciel.',
        ParagraphStyle('EndNote', fontName='Helvetica', fontSize=10, textColor=rl_colors.HexColor('#94a3b8'), alignment=TA_CENTER)
    ))

    doc.build(story)
    buffer.seek(0)
    return FileResponse(
        buffer, as_attachment=True,
        filename=f'manuel-utilisation-{date.today().isoformat()}.pdf'
    )


# ===== MANUEL D'UTILISATION HTML =====
def user_manual(request):
    return render(request, 'store/user_manual.html')
